#Zoho Final
from django.shortcuts import render,redirect
from Register_Login.models import *
from Register_Login.views import logout
from django.contrib import messages
from django.conf import settings
from datetime import date
from datetime import datetime, timedelta
from Company_Staff.models import *
from django.db import models
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.core.mail import send_mail
from django.core.mail import EmailMessage
from xhtml2pdf import pisa
from django.template.loader import get_template
from bs4 import BeautifulSoup
import io,os
import csv
import math
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from django.http import HttpResponse,HttpResponseRedirect
from io import BytesIO
from django.db.models import Max
from django.db.models import Q
from django.http import JsonResponse,HttpResponse,HttpResponseRedirect
from django.urls import reverse
from django.shortcuts import render,redirect,get_object_or_404
from . models import *
from decimal import Decimal
from .models import Journal, JournalEntry, LoginDetails, CompanyDetails, StaffDetails,JournalTransactionHistory,JournalComment
import openpyxl
from openpyxl import Workbook, load_workbook
import datetime

# Create your views here.




# -------------------------------Company section--------------------------------

# company dashboard
def company_dashboard(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')

        # Calculate the date 20 days before the end date for payment term renew
        reminder_date = dash_details.End_date - timedelta(days=20)
        current_date = date.today()
        alert_message = current_date >= reminder_date
        
        payment_request = True if PaymentTermsUpdates.objects.filter(company=dash_details,update_action=1,status='Pending').exists() else False

        # Calculate the number of days between the reminder date and end date
        days_left = (dash_details.End_date - current_date).days
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'alert_message':alert_message,
            'days_left':days_left,
            'payment_request':payment_request,
        }
        return render(request, 'company/company_dash.html', context)
    else:
        return redirect('/')


# company staff request for login approval
def company_staff_request(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        staff_request=StaffDetails.objects.filter(company=dash_details.id, company_approval=0).order_by('-id')
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'requests':staff_request,
        }
        return render(request, 'company/staff_request.html', context)
    else:
        return redirect('/')

# company staff accept or reject
def staff_request_accept(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    staff.company_approval=1
    staff.save()
    return redirect('company_staff_request')

def staff_request_reject(request,pk):
    staff=StaffDetails.objects.get(id=pk)
    login_details=LoginDetails.objects.get(id=staff.company.id)
    login_details.delete()
    staff.delete()
    return redirect('company_staff_request')


# All company staff view, cancel staff approval
def company_all_staff(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        all_staffs=StaffDetails.objects.filter(company=dash_details.id, company_approval=1).order_by('-id')
       
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'staffs':all_staffs,
        }
        return render(request, 'company/all_staff_view.html', context)
    else:
        return redirect('/')

def staff_approval_cancel(request, pk):
    """
    Sets the company approval status to 2 for the specified staff member, effectively canceling staff approval.

    This function is designed to be used for canceling staff approval, and the company approval value is set to 2.
    This can be useful for identifying resigned staff under the company in the future.

    """
    staff = StaffDetails.objects.get(id=pk)
    staff.company_approval = 2
    staff.save()
    return redirect('company_all_staff')


# company profile, profile edit
def company_profile(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        terms=PaymentTerms.objects.all()
        payment_history=dash_details.previous_plans.all()

        # Calculate the date 20 days before the end date
        reminder_date = dash_details.End_date - timedelta(days=20)
        current_date = date.today()
        renew_button = current_date >= reminder_date

        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'renew_button': renew_button,
            'terms':terms,
            'payment_history':payment_history,
        }
        return render(request, 'company/company_profile.html', context)
    else:
        return redirect('/')

def company_profile_editpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'company/company_profile_editpage.html', context)
    else:
        return redirect('/')

def company_profile_basicdetails_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        if request.method == 'POST':
            # Get data from the form
            log_details.first_name = request.POST.get('fname')
            log_details.last_name = request.POST.get('lname')
            log_details.email = request.POST.get('eid')
            log_details.username = request.POST.get('uname')
            log_details.save()
            messages.success(request,'Updated')
            return redirect('company_profile_editpage') 
        else:
            return redirect('company_profile_editpage') 

    else:
        return redirect('/')
    
def company_password_change(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        if request.method == 'POST':
            # Get data from the form
            password = request.POST.get('pass')
            cpassword = request.POST.get('cpass')
            if password == cpassword:
                if LoginDetails.objects.filter(password=password).exists():
                    messages.error(request,'Use another password')
                    return redirect('company_profile_editpage')
                else:
                    log_details.password=password
                    log_details.save()

            messages.success(request,'Password Changed')
            return redirect('company_profile_editpage') 
        else:
            return redirect('company_profile_editpage') 

    else:
        return redirect('/')
       
def company_profile_companydetails_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details = LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

        if request.method == 'POST':
            # Get data from the form
            gstno = request.POST.get('gstno')
            profile_pic = request.FILES.get('image')

            # Update the CompanyDetails object with form data
            dash_details.company_name = request.POST.get('cname')
            dash_details.contact = request.POST.get('phone')
            dash_details.address = request.POST.get('address')
            dash_details.city = request.POST.get('city')
            dash_details.state = request.POST.get('state')
            dash_details.country = request.POST.get('country')
            dash_details.pincode = request.POST.get('pincode')
            dash_details.pan_number = request.POST.get('pannumber')

            if gstno:
                dash_details.gst_no = gstno

            if profile_pic:
                dash_details.profile_pic = profile_pic

            dash_details.save()

            messages.success(request, 'Updated')
            return redirect('company_profile_editpage')
        else:
            return redirect('company_profile_editpage')
    else:
        return redirect('/')    

# company modules editpage
def company_module_editpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'company/company_module_editpage.html', context)
    else:
        return redirect('/')

def company_module_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')

        # Check for any previous module update request
        if ZohoModules.objects.filter(company=dash_details,status='Pending').exists():
            messages.warning(request,'You have a pending update request, wait for approval or contact our support team for any help..?')
            return redirect('company_profile')
        if request.method == 'POST':
            # Retrieve values
            items = request.POST.get('items', 0)
            price_list = request.POST.get('price_list', 0)
            stock_adjustment = request.POST.get('stock_adjustment', 0)
            godown = request.POST.get('godown', 0)

            cash_in_hand = request.POST.get('cash_in_hand', 0)
            offline_banking = request.POST.get('offline_banking', 0)
            upi = request.POST.get('upi', 0)
            bank_holders = request.POST.get('bank_holders', 0)
            cheque = request.POST.get('cheque', 0)
            loan_account = request.POST.get('loan_account', 0)

            customers = request.POST.get('customers', 0)
            invoice = request.POST.get('invoice', 0)
            estimate = request.POST.get('estimate', 0)
            sales_order = request.POST.get('sales_order', 0)
            recurring_invoice = request.POST.get('recurring_invoice', 0)
            retainer_invoice = request.POST.get('retainer_invoice', 0)
            credit_note = request.POST.get('credit_note', 0)
            payment_received = request.POST.get('payment_received', 0)
            delivery_challan = request.POST.get('delivery_challan', 0)

            vendors = request.POST.get('vendors', 0)
            bills = request.POST.get('bills', 0)
            recurring_bills = request.POST.get('recurring_bills', 0)
            vendor_credit = request.POST.get('vendor_credit', 0)
            purchase_order = request.POST.get('purchase_order', 0)
            expenses = request.POST.get('expenses', 0)
            recurring_expenses = request.POST.get('recurring_expenses', 0)
            payment_made = request.POST.get('payment_made', 0)

            projects = request.POST.get('projects', 0)

            chart_of_accounts = request.POST.get('chart_of_accounts', 0)
            manual_journal = request.POST.get('manual_journal', 0)

            eway_bill = request.POST.get('ewaybill', 0)

            employees = request.POST.get('employees', 0)
            employees_loan = request.POST.get('employees_loan', 0)
            holiday = request.POST.get('holiday', 0)
            attendance = request.POST.get('attendance', 0)
            salary_details = request.POST.get('salary_details', 0)

            reports = request.POST.get('reports', 0)

            update_action=1
            status='Pending'

            # Create a new ZohoModules instance and save it to the database
            data = ZohoModules(
                company=dash_details,
                items=items, price_list=price_list, stock_adjustment=stock_adjustment, godown=godown,
                cash_in_hand=cash_in_hand, offline_banking=offline_banking, upi=upi, bank_holders=bank_holders,
                cheque=cheque, loan_account=loan_account,
                customers=customers, invoice=invoice, estimate=estimate, sales_order=sales_order,
                recurring_invoice=recurring_invoice, retainer_invoice=retainer_invoice, credit_note=credit_note,
                payment_received=payment_received, delivery_challan=delivery_challan,
                vendors=vendors, bills=bills, recurring_bills=recurring_bills, vendor_credit=vendor_credit,
                purchase_order=purchase_order, expenses=expenses, recurring_expenses=recurring_expenses,
                payment_made=payment_made,
                projects=projects,
                chart_of_accounts=chart_of_accounts, manual_journal=manual_journal,
                eway_bill=eway_bill,
                employees=employees, employees_loan=employees_loan, holiday=holiday,
                attendance=attendance, salary_details=salary_details,
                reports=reports,update_action=update_action,status=status    
            )
            data.save()
            messages.success(request,"Request sent successfully. Please wait for approval.")
            return redirect('company_profile')
        else:
            return redirect('company_module_editpage')  
    else:
        return redirect('/')


def company_renew_terms(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

        # Check for any previous  extension request
        if PaymentTermsUpdates.objects.filter(company=dash_details,update_action=1,status='Pending').exists():
            messages.warning(request,'You have a pending request, wait for approval or contact our support team for any help..?')
            return redirect('company_profile')
        if request.method == 'POST':
            select=request.POST['select']
            terms=PaymentTerms.objects.get(id=select)
            update_action=1
            status='Pending'
            newterms=PaymentTermsUpdates(
               company=dash_details,
               payment_term=terms,
               update_action=update_action,
               status=status 
            )
            newterms.save()
            messages.success(request,'Request sent successfully, Please wait for approval...')
            return redirect('company_profile')
        else:
            return redirect('company_profile')
    else:
        return redirect('/')

# company notifications and messages
def company_notifications(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        notifications = dash_details.notifications.filter(is_read=0).order_by('-date_created','-time')
        end_date = dash_details.End_date
        company_days_remaining = (end_date - date.today()).days
        payment_request = True if PaymentTermsUpdates.objects.filter(company=dash_details,update_action=1,status='Pending').exists() else False
        
        print(company_days_remaining)
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'notifications':notifications,
            'days_remaining':company_days_remaining,
            'payment_request':payment_request,
        }

        return render(request,'company/company_notifications.html',context)
        
    else:
        return redirect('/')
        
        
def company_message_read(request,pk):
    '''
    message read functions set the is_read to 1, 
    by default it is 0 means not seen by user.

    '''
    notification=Notifications.objects.get(id=pk)
    notification.is_read=1
    notification.save()
    return redirect('company_notifications')
    
    
def company_payment_history(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/') 
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        payment_history=dash_details.previous_plans.all()

        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'payment_history':payment_history,
            
        }
        return render(request,'company/company_payment_history.html', context)
    else:
        return redirect('/')
        
def company_trial_feedback(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/') 
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        trial_instance = TrialPeriod.objects.get(company=dash_details)
        if request.method == 'POST':
            interested = request.POST.get('interested')
            feedback=request.POST.get('feedback') 
            
            trial_instance.interested_in_buying=1 if interested =='yes' else 2
            trial_instance.feedback=feedback
            trial_instance.save()

            if interested =='yes':
                return redirect('company_profile')
            else:
                return redirect('company_dashboard')
        else:
            return redirect('company_dashboard')
    else:
        return redirect('/')
# -------------------------------Staff section--------------------------------

# staff dashboard
def staff_dashboard(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context={
            'details':dash_details,
            'allmodules': allmodules,
        }
        return render(request,'staff/staff_dash.html',context)
    else:
        return redirect('/')


# staff profile
def staff_profile(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context={
            'details':dash_details,
            'allmodules': allmodules,
        }
        return render(request,'staff/staff_profile.html',context)
    else:
        return redirect('/')


def staff_profile_editpage(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        context = {
            'details': dash_details,
            'allmodules': allmodules
        }
        return render(request, 'staff/staff_profile_editpage.html', context)
    else:
        return redirect('/')

def staff_profile_details_edit(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        dash_details = StaffDetails.objects.get(login_details=log_details,company_approval=1)
        if request.method == 'POST':
            # Get data from the form
            log_details.first_name = request.POST.get('fname')
            log_details.last_name = request.POST.get('lname')
            log_details.email = request.POST.get('eid')
            log_details.username = request.POST.get('uname')
            log_details.save()
            dash_details.contact = request.POST.get('phone')
            old=dash_details.image
            new=request.FILES.get('profile_pic')
            print(new,old)
            if old!=None and new==None:
                dash_details.image=old
            else:
                print(new)
                dash_details.image=new
            dash_details.save()
            messages.success(request,'Updated')
            return redirect('staff_profile_editpage') 
        else:
            return redirect('staff_profile_editpage') 

    else:
        return redirect('/')

def staff_password_change(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details= LoginDetails.objects.get(id=log_id)
        if request.method == 'POST':
            # Get data from the form
            password = request.POST.get('pass')
            cpassword = request.POST.get('cpass')
            if password == cpassword:
                if LoginDetails.objects.filter(password=password).exists():
                    messages.error(request,'Use another password')
                    return redirect('staff_profile_editpage')
                else:
                    log_details.password=password
                    log_details.save()

            messages.success(request,'Password Changed')
            return redirect('staff_profile_editpage') 
        else:
            return redirect('staff_profile_editpage') 

    else:
        return redirect('/')


    
def company_gsttype_change(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details = LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

        if request.method == 'POST':
            # Get data from the form
            
            gstno = request.POST.get('gstno')
            gsttype = request.POST.get('gsttype')

            # Check if gsttype is one of the specified values
            if gsttype in ['unregistered Business', 'Overseas', 'Consumer']:
                dash_details.gst_no = None
            else:
                if gstno:
                    dash_details.gst_no = gstno
                else:
                    messages.error(request,'GST Number is not entered*')
                    return redirect('company_profile_editpage')


            dash_details.gst_type = gsttype

            dash_details.save()
            messages.success(request,'GST Type changed')
            return redirect('company_profile_editpage')
        else:
            return redirect('company_profile_editpage')
    else:
        return redirect('/') 
    

# -------------------------------Zoho Modules section--------------------------------
# -------------------------------Zoho Modules section--------------------------------

#--------------------------------------------------- TINTO VIEW ITEMS START-------------------------------------------

# items llist
    
def items_list(request):                                                                
     if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=login_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                item=Items.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'item':item,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/items/items_list.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            item=Items.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            context = {
                    'details': dash_details,
                    'item': item,
                    'allmodules': allmodules,
            }
        return render(request,'zohomodules/items/items_list.html',context)

   
   
# create Items

def new_items(request):                                                              
    if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=login_id)
    if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                item=Items.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                units = Unit.objects.filter(company=dash_details.company)
                accounts=Chart_of_Accounts.objects.filter(company=dash_details.company)
                context = {
                     'details': dash_details,
                    'units': units,
                    'allmodules': allmodules,
                    'accounts':accounts
                }
                return render(request,'zohomodules/items/newitem.html',context)
    if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            item=Items.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            units = Unit.objects.filter(company=dash_details)
            accounts=Chart_of_Accounts.objects.filter(company=dash_details)
            context = {
                    'details': dash_details,
                    'units': units,
                    'allmodules': allmodules,
                    'accounts':accounts
            }
    
            return render(request, 'zohomodules/items/newitem.html',context)
# create Items
def create_item(request):                                                                #new by tinto mt
    
    login_id = request.session['login_id']
    if 'login_id' not in request.session:
        return redirect('/')
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.Date=date.today()
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            uid=Unit.objects.get(id=unit_id)
            # unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = uid
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            # track = request.POST.get("trackState",None)
            track_state_value = request.POST.get("trackstate", None)

# Check if the checkbox is checked
            if track_state_value == "on":
                a.track_inventory = 1
            else:
                a.track_inventory = 0

            
            minstock=request.POST.get("minimum_stock",None)
            if minstock != "":
                a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                a.minimum_stock_to_maintain = 0
            a.activation_tag = 'Active'
            a.type = 'Opening Stock'
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
            a.current_stock=request.POST.get("openstock",None)
            a.opening_stock_per_unit = request.POST.get("rate",None)
            item_name= request.POST.get("name",None)
            hsncode=request.POST.get("hsn",None)
            
            if Items.objects.filter(item_name=item_name, company=c).exists():
                error='yes'
                messages.error(request,'Item with same name exsits !!!')
                return redirect('new_items')
            elif Items.objects.filter(hsn_code=hsncode, company=c).exists():
                error='yes'
                messages.error(request,'Item with same  hsn code exsits !!!')
                return redirect('new_items')
            else:
                a.save()    
                t=Items.objects.get(id=a.id)
                b.items=t
                b.save()
                return redirect('items_list')
    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
            a.item_type = request.POST.get("type",None)
            a.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            unit_instance = get_object_or_404(Unit, id=unit_id)
            a.unit = unit_instance
            a.hsn_code = request.POST.get("hsn",None)
            a.tax_reference = request.POST.get("radio",None)
            a.intrastate_tax = request.POST.get("intra",None)
            a.interstate_tax= request.POST.get("inter",None)
            a.selling_price = request.POST.get("sel_price",None)
            a.sales_account = request.POST.get("sel_acc",None)
            a.sales_description = request.POST.get("sel_desc",None)
            a.purchase_price = request.POST.get("cost_price",None)
            a.purchase_account = request.POST.get("cost_acc",None)
            a.purchase_description = request.POST.get("pur_desc",None)
            # track_state_value = request.POST.get("trackState", None)

            track_state_value = request.POST.get("trackstate", None)

            # Check if the checkbox is checked
            if track_state_value == "on":
                a.track_inventory = 1
            else:
                a.track_inventory = 0
            minstock=request.POST.get("minimum_stock",None)
            item_name= request.POST.get("name",None)
            hsncode=request.POST.get("hsn",None)
            
            if minstock != "":
                a.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                a.minimum_stock_to_maintain = 0
            # a.activation_tag = request.POST.get("status",None)
            a.inventory_account = request.POST.get("invacc",None)
            a.opening_stock = request.POST.get("openstock",None)
            a.current_stock=request.POST.get("openstock",None)
        
        

        
            if Items.objects.filter(item_name=item_name,company=c).exists():
                error='yes'
                messages.error(request,'Item with same name exsits !!!')
                return redirect('new_items')
            elif Items.objects.filter(hsn_code=hsncode, company=c).exists():
                error='yes'
                messages.error(request,'Item with same  hsn code exsits !!!')
                return redirect('new_items')
            else:
                a.save()    
                t=Items.objects.get(id=a.id)
                b.items=t
                b.save()
                return redirect('items_list')
    return redirect('items_list')

# create unit
def add_unit(request):                                                                #new by tinto mt (item)
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)

    if log_user.user_type == 'Company':
        if request.method == 'POST':
            c = CompanyDetails.objects.get(login_details=login_id)
            unit_name = request.POST['units']
            
            if Unit.objects.filter(unit_name=unit_name, company=c).exists():
                return JsonResponse({"message": "error"})
            else:
                unit = Unit(unit_name=unit_name, company=c)  
                unit.save()  
                return JsonResponse({"message": "success"})

    elif log_user.user_type == 'Staff':
        if request.method == 'POST':
            staff = LoginDetails.objects.get(id=login_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c = sf.company
            unit_name = request.POST['units']
            
            if Unit.objects.filter(unit_name=unit_name, company=c).exists():
                return JsonResponse({"message": "error"})
            else:
                unit = Unit(unit_name=unit_name, company=c)  
                unit.save()  
                return JsonResponse({"message": "success"})

    return JsonResponse({"message": "success"})
# create unit


    
def unit_dropdown(request):                                                               
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_user)
            options = {}
            option_objects = Unit.objects.filter(company=dash_details)
            for option in option_objects:
                unit_name=option.unit_name
            options[option.id] = [unit_name,f"{unit_name}"]
            return JsonResponse(options)
      

    elif log_user.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_user)
            options = {}
            option_objects = Unit.objects.filter(company=dash_details.company)
            for option in option_objects:
                unit_name=option.unit_name
            options[option.id] = [unit_name,f"{unit_name}"]
            return JsonResponse(options)
             



def add_account(request):                                                              
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        if request.method == 'POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.logindetails=log_user
            b.action="Created"
            b.Date=date.today()
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.description = request.POST.get("description",None)
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name, company=c).exists():
                return JsonResponse({"message": "error"})
            else:
          
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                acc_id = a.id  
                acc_name=a.account_name
                response_data = {
                "message": "success",
                "acc_id":acc_id,
                "acc_name":acc_name,
        
                         }

                return JsonResponse(response_data)
        

    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            a=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.description = request.POST.get("description",None)
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name, company=c).exists():
                return JsonResponse({"message": "error"})
            else:
          
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                acc_id = a.id  
                acc_name=a.account_name
                response_data = {
                "message": "success",
                "acc_id":acc_id,
                "acc_name":acc_name,
        
                         }

                return JsonResponse(response_data)
        
      
        
    return redirect('newitems')

def account_dropdown(request):                                                                
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_user)
            options = {}
            option_objects = Chart_of_Accounts.objects.filter(Q(company=dash_details) & (Q(account_type='Expense') | Q(account_type='Other Expense') | Q(account_type='Cost Of Goods Sold')))
            for option in option_objects:
                account_name=option.account_name
                account_type=option.account_type
                options[option.id] = [account_name,f"{account_name}"]
            return JsonResponse(options)
    elif log_user.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_user)
            options = {}
       
            option_objects = Chart_of_Accounts.objects.filter(Q(company=dash_details.company) & (Q(account_type='Expense') | Q(account_type='Other Expense') | Q(account_type='Cost Of Goods Sold')))
            for option in option_objects:
                account_name=option.account_name
                options[option.id] = [account_name,f"{account_name}"]
            return JsonResponse(options)


    
    
def itemsoverview(request,pk):                                                                
    if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=login_id)
    if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                item=Items.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
              
                items=Items.objects.filter(company=dash_details.company)
                selitem=Items.objects.get(id=pk)
                est_comments=Items_comments.objects.filter(Items=pk)
                stock_value=selitem.opening_stock*selitem.purchase_price  
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
                context = {
                     'details': dash_details,
                
                    'allmodules': allmodules,
                    'items':items,
                    'selitem':selitem,
                    'stock_value':stock_value,
                    'latest_item_id':filtered_data,
                    'est_comments':est_comments
                }
                return render(request, 'zohomodules/items/itemsoverview.html',context)
    if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
       
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            items=Items.objects.filter(company=dash_details)
            selitem=Items.objects.get(id=pk)
            est_comments=Items_comments.objects.filter(Items=pk)
            stock_value=selitem.opening_stock*selitem.purchase_price  
            latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
            filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
            context = {
                    'details': dash_details,
                   
                    'allmodules': allmodules,
                    'items':items,
                    'selitem':selitem,
                    'stock_value':stock_value,
                    'latest_item_id':filtered_data,
                    'est_comments':est_comments
            }
    
            return render(request, 'zohomodules/items/itemsoverview.html',context)


    return render(request, 'zohomodules/items/itemsoverview.html')


def edititems(request, pr):                                                                #new by tinto mt
    if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    
    # Retrieve the chart of accounts entry
    item = get_object_or_404(Items, id=pr)
    

    # Check if 'company_id' is in the session

    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
      
     
        dash_details = CompanyDetails.objects.get(login_details=log_user)
        units = Unit.objects.filter(company=dash_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        item = get_object_or_404(Items, id=pr)
        accounts=Chart_of_Accounts.objects.filter(company=dash_details)
        units = Unit.objects.filter(company=dash_details)
        context = {
                    'item': item,
                    'units':units,
                    'details': dash_details,
                   'accounts': accounts,
                    'allmodules': allmodules,
            }
       
    
        
        if request.method=='POST':
   
            b=Item_Transaction_History()
            # c = CompanyDetails.objects.get(login_details=company_id)
            b.company=dash_details
            b.logindetails=log_user
            b.action="Edited"
            b.Date=date.today()
            item.login_details=log_user
            item.company=dash_details
            item.item_type = request.POST.get("type",None)
            item.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            unit_instance = get_object_or_404(Unit, id=unit_id)
            item.unit = unit_instance
            item.hsn_code = request.POST.get("hsn",None)
            item.tax_reference = request.POST.get("radio",None)
            if request.POST.get("radio",None) == 'taxable':

                item.intrastate_tax = request.POST.get("intra",None)
                item.interstate_tax= request.POST.get("inter",None)
            elif request.POST.get("radio",None) == 'None-Taxable':
                item.intrastate_tax = 0
                item.interstate_tax= 0
            item.selling_price = request.POST.get("sel_price",None)
            item.sales_account = request.POST.get("sel_acc",None)
            item.sales_description = request.POST.get("sel_desc",None)
            item.purchase_price = request.POST.get("cost_price",None)
            item.purchase_account = request.POST.get("cost_acc",None)
            item.purchase_description = request.POST.get("pur_desc",None)
            minstock=request.POST.get("minimum_stock",None)
            if minstock != "":
                item.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                item.minimum_stock_to_maintain = 0
            # item.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            # item.activation_tag = request.POST.get("status",None)
            item.inventory_account = request.POST.get("invacc",None)
            item.opening_stock = request.POST.get("openstock",None)
            item.opening_stock_per_unit = request.POST.get("rate",None)
            item.current_stock= request.POST.get("openstock",None)
            track_state_value = request.POST.get("trackstate", None)
            if track_state_value == "on":
                item.track_inventory = 1
            else:
                item.track_inventory = 0
            
            # Save the changes
            item.save()
            t=Items.objects.get(id=item.id)
            b.items=t
            b.save()
            # Redirect to another page after successful update
            return redirect('itemsoverview', pr)
        return render(request, 'zohomodules/items/edititems.html',context)
    if log_user.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_user)
                
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        item = get_object_or_404(Items, id=pr)
        units = Unit.objects.filter(company=dash_details.company)
        accounts=Chart_of_Accounts.objects.filter(company=dash_details.company)
        context = {
                    'item': item,
                    'units':units,
                    'details': dash_details,
                    'accounts': accounts,
                   
                    'allmodules': allmodules,
            }
 
        if request.method=='POST':
            a=Items()
            b=Item_Transaction_History()

            c=dash_details.company
            b.company=c
            b.logindetails=log_user
            b.action="Edited"
            b.Date=date.today()
            a.login_details=log_user
            a.company=c
            item.item_type = request.POST.get("type",None)
            item.item_name = request.POST.get("name",None)
            unit_id = request.POST.get("unit")
            unit_instance = get_object_or_404(Unit, id=unit_id)
            item.unit = unit_instance
            item.hsn_code = request.POST.get("hsn",None)
            item.tax_reference = request.POST.get("radio",None)
            item.intrastate_tax = request.POST.get("intra",None)
            item.interstate_tax= request.POST.get("inter",None)
            item.selling_price = request.POST.get("sel_price",None)
            item.sales_account = request.POST.get("sel_acc",None)
            item.sales_description = request.POST.get("sel_desc",None)
            item.purchase_price = request.POST.get("cost_price",None)
            item.purchase_account = request.POST.get("cost_acc",None)
            item.purchase_description = request.POST.get("pur_desc",None)
            minstock=request.POST.get("minimum_stock",None)
            if minstock != "":
                item.minimum_stock_to_maintain = request.POST.get("minimum_stock",None)
            else:
                item.minimum_stock_to_maintain = 0
            # item.activation_tag = request.POST.get("status",None)
            item.inventory_account = request.POST.get("invacc",None)
            item.opening_stock = request.POST.get("openstock",None)
            item.current_stock= request.POST.get("openstock",None)
            item.opening_stock_per_unit = request.POST.get("rate",None)
            track_state_value = request.POST.get("trackstate", None)
            if track_state_value == "on":
                item.track_inventory = 1
            else:
                item.track_inventory = 0
            item.save()
            t=Items.objects.get(id=item.id)
            b.items=t
            b.save()

            return redirect('itemsoverview', pr)
 
        return render(request, 'zohomodules/items/edititems.html', context)
   
def item_status_edit(request, pv):                                                                #new by tinto mt
    
    selitem = Items.objects.get(id=pv)

    if selitem.activation_tag == 'Active':
        selitem.activation_tag = 'inactive'
        selitem.save()
    elif selitem.activation_tag != 'Active':
        selitem.activation_tag = 'Active'
        selitem.save()

    selitem.save()

    return redirect('itemsoverview',pv)


def shareItemToEmail(request,pt):                                                                #new by tinto mt
    if request.user: 
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']
                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                print(emails_list)
                print('1')
           
           
                item = Items.objects.get(id=pt)
                context = {
                
                    'selitem':item,
                }
                print('2')
                template_path = 'zohomodules/items/itememailpdf.html'
                print('3')
                template = get_template(template_path)
                print('4')
                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)#, link_callback=fetch_resources)
                pdf = result.getvalue()
                print('5')
                filename = f'Item Transactions.pdf'
                subject = f"Transactipns"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached Item transactions. \n{email_message}\n\n--\nRegards,\n{item.item_name}\n{item.item_type}", from_email=settings.EMAIL_HOST_USER,to=emails_list)
                email.attach(filename,pdf,"application/pdf")
                email.send(fail_silently=False)
                msg = messages.success(request, 'Details has been shared via email successfully..!')
                return redirect(itemsoverview,pt)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(itemsoverview,pt)   
        
def deleteitem(request,pl):                                                                #new by tinto mt
    items=Items.objects.filter(id=pl)
    items.delete()
    
    return redirect(items_list)

def delete_item_comment(request,ph,pr):                                                                #new by tinto mt
    items=Items_comments.objects.filter(id=ph)
    items.delete()
    ac=Items.objects.get(id=pr)
    
    return redirect(itemsoverview,ac.id)


def add_item_comment(request,pc):                                                                #new by tinto mt

    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        if request.method=="POST":
                    
                    com=Items_comments()
                    c = CompanyDetails.objects.get(login_details=company_id)
            
                    comment_comments=request.POST['comment']
                    com.company=c
                    com.logindetails=log_user
                    com.comments=comment_comments
                    item=Items.objects.get(id=pc)
                    com.Items=item
                    
                    com.save()
                    return redirect('itemsoverview',pc)

    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            com=Items_comments()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            
            comment_comments=request.POST['comment']
            com.company=c
            com.logindetails=log_user
            com.comments=comment_comments
            item=Items.objects.get(id=pc)
            com.Items=item
                    
            com.save()
            return redirect('itemsoverview',pc)
    return redirect('itemsoverview',pc)
        




         
def downloadItemSampleImportFile(request):                                                                  #new by tinto mt
    estimate_table_data = [['No.','ITEM TYPE','ITEM NAME','HSN','TAX REFERENCE','INTRASTATE TAX','INTERSTATE TAX','SELLING PRICE','SALES ACCOUNT','SALES DESCRIPTION','PURCHASE PRICE','PURCHASE ACCOUNT','PURCHASE DESCRIPTION','MINIMUM STOCK TO MAINTAIN','ACTIVATION TAG','OPENING STOCK','CURRENT STOCK','OPENING STOCK PER UNIT']]      
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = 'Sheet1'
    

    # Populate the sheets with data
    for row in estimate_table_data:
        sheet1.append(row)  
    
    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=expense_sample_file.xlsx'
     # Save the workbook to the response
    wb.save(response)
    return response





def import_item(request):                                                                #new by tinto mt
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)

    if log_user.user_type == 'Company':
        company_id = request.session['login_id']

        if request.method == 'POST' and 'excel_file' in request.FILES:
            company = CompanyDetails.objects.get(login_details=company_id)
            excel_file = request.FILES.get('excel_file')
            wb = load_workbook(excel_file)

            try:
                ws = wb["Sheet1"]
                header_row = ws[1]
                column_names = [cell.value for cell in header_row]
                print("Column Names:", column_names)
            except KeyError:
                print('Sheet not found')
                messages.error(request, '`Sheet1` not found in the Excel file. Please check.')
                return redirect('expensepage')

            expected_columns = ['No.', 'ITEM TYPE', 'ITEM NAME', 'HSN', 'TAX REFERENCE', 'INTRASTATE TAX', 'INTERSTATE TAX',
                                'SELLING PRICE', 'SALES ACCOUNT', 'SALES DESCRIPTION', 'PURCHASE PRICE',
                                'PURCHASE ACCOUNT', 'PURCHASE DESCRIPTION', 'MINIMUM STOCK TO MAINTAIN', 'ACTIVATION TAG',
                                'OPENING STOCK', 'CURRENT STOCK', 'OPENING STOCK PER UNIT']

            if column_names != expected_columns:
                print('Invalid sheet columns or order')
                messages.error(request, 'Sheet column names or order is not in the required format. Please check.')
                return redirect("comapny_items")

            for row in ws.iter_rows(min_row=2, values_only=True):
                _, item_type, item_name, hsn, tax_reference, intrastate_tax, interstate_tax, selling_price, sales_account, \
                sales_description, purchase_price, purchase_account, purchase_description, min_stock, activation_tag, \
                opening_stock, current_stock, opening_stock_per_unit = row

                # Fetching the 'Unit' instance with id=1 (you may adjust this based on your 'Unit' model)
                unit_instance = Unit.objects.get(pk=1)

                # Creating an instance of the 'Items' model and saving it
                item = Items(
                    login_details=log_user,
                    company=company,
                    unit=unit_instance,  # Use the fetched 'Unit' instance
                    item_type=item_type,
                    item_name=item_name,
                    hsn_code=hsn,
                    tax_reference=tax_reference,
                    intrastate_tax=intrastate_tax,
                    interstate_tax=interstate_tax,
                    selling_price=selling_price,
                    sales_account=sales_account,
                    sales_description=sales_description,
                    purchase_price=purchase_price,
                    purchase_account=purchase_account,
                    purchase_description=purchase_description,
                    minimum_stock_to_maintain=min_stock,
                    activation_tag=activation_tag,
                    inventory_account="Inventory Account",
                    opening_stock=opening_stock,
                    opening_stock_per_unit=opening_stock_per_unit
                )
                item.save()

            messages.success(request, 'Data imported successfully!')
            return redirect("items_list")
        else:
            messages.error(request, 'Invalid request. Please check the file and try again.')
            return redirect("items_list")
    else:
        messages.error(request, 'Invalid user type. Please check your user type.')
        return redirect("items_list")


def item_view_sort_by_name(request, pk):    
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
            
                items = list(Items.objects.filter(company=dash_details.company).values())

    # Sort the items by the 'item_name' field
                sorted_items = sorted(items, key=lambda r: r['item_name'])

                # Get the selected item by ID
                selitem = Items.objects.get(id=pk)

                # Fetch related comments for the selected item
                est_comments = Items_comments.objects.filter(Items=pk)

                # Calculate stock value for the selected item
                stock_value = selitem.opening_stock * selitem.purchase_price

                # Find the latest date for the item transaction history
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']

                # Filter transaction history for the latest date and the selected item
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)

                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'items': sorted_items, 
                    'selitem': selitem, 
                    'stock_value': stock_value, 
                    'latest_item_id': filtered_data, 
                    'est_comments': est_comments
                    
            } 
                return render(request,'zohomodules/items/itemsoverview.html',content)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
                items = list(Items.objects.filter(company=dash_details).values())

    # Sort the items by the 'item_name' field
                sorted_items = sorted(items, key=lambda r: r['item_name'])

                # Get the selected item by ID
                selitem = Items.objects.get(id=pk)

                # Fetch related comments for the selected item
                est_comments = Items_comments.objects.filter(Items=pk)

                # Calculate stock value for the selected item
                stock_value = selitem.opening_stock * selitem.purchase_price

                # Find the latest date for the item transaction history
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']

                # Filter transaction history for the latest date and the selected item
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)

                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                content = {
                        'details': dash_details,
                        'allmodules': allmodules,
                        'items': sorted_items, 
                        'selitem': selitem, 
                        'stock_value': stock_value, 
                        'latest_item_id': filtered_data, 
                        'est_comments': est_comments
                        
                }  
                return render(request,'zohomodules/items/itemsoverview.html',content)

def item_view_sort_by_hsn(request, pk):      
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
            
                items = list(Items.objects.filter(company=dash_details.company).values())

    # Sort the items by the 'item_name' field
                sorted_items = sorted(items, key=lambda r: r['hsn_code'])

                # Get the selected item by ID
                selitem = Items.objects.get(id=pk)

                # Fetch related comments for the selected item
                est_comments = Items_comments.objects.filter(Items=pk)

                # Calculate stock value for the selected item
                stock_value = selitem.opening_stock * selitem.purchase_price

                # Find the latest date for the item transaction history
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']

                # Filter transaction history for the latest date and the selected item
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)

                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'items': sorted_items, 
                    'selitem': selitem, 
                    'stock_value': stock_value, 
                    'latest_item_id': filtered_data, 
                    'est_comments': est_comments
                    
            } 
                return render(request,'zohomodules/items/itemsoverview.html',content)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
                items = list(Items.objects.filter(company=dash_details).values())

    # Sort the items by the 'item_name' field
                sorted_items = sorted(items, key=lambda r: r['item_name'])

                # Get the selected item by ID
                selitem = Items.objects.get(id=pk)

                # Fetch related comments for the selected item
                est_comments = Items_comments.objects.filter(Items=pk)

                # Calculate stock value for the selected item
                stock_value = selitem.opening_stock * selitem.purchase_price

                # Find the latest date for the item transaction history
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']

                # Filter transaction history for the latest date and the selected item
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)

                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                content = {
                        'details': dash_details,
                        'allmodules': allmodules,
                        'items': sorted_items, 
                        'selitem': selitem, 
                        'stock_value': stock_value, 
                        'latest_item_id': filtered_data, 
                        'est_comments': est_comments
                        
                }  
                return render(request,'zohomodules/items/itemsoverview.html',content)

def filter_item_view_Active(request,pk):          
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
            
           

    # Sort the items by the 'item_name' field
                items=Items.objects.filter(activation_tag='Active',company=dash_details.company)  
                selitem=Items.objects.get(id=pk)
                est_comments=Items_comments.objects.filter(Items=pk)
                stock_value=selitem.opening_stock*selitem.purchase_price  
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'items':items,
                    'selitem':selitem,
                    'stock_value':stock_value,
                    'latest_item_id':filtered_data,
                    'est_comments':est_comments
                    
            } 
                return render(request,'zohomodules/items/itemsoverview.html',content)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
         

    # Sort the items by the 'item_name' field
                items=Items.objects.filter(activation_tag='Active',company=dash_details)  
                selitem=Items.objects.get(id=pk)
                est_comments=Items_comments.objects.filter(Items=pk)
                stock_value=selitem.opening_stock*selitem.purchase_price  
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                content = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'items':items,
                    'selitem':selitem,
                    'stock_value':stock_value,
                    'latest_item_id':filtered_data,
                    'est_comments':est_comments
                    
            } 
                return render(request,'zohomodules/items/itemsoverview.html',content) 

def filter_item_view_inActive(request,pk):         
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
            
           

    # Sort the items by the 'item_name' field
                items=Items.objects.filter(activation_tag='inactive',company=dash_details.company)  
                selitem=Items.objects.get(id=pk)
                est_comments=Items_comments.objects.filter(Items=pk)
                stock_value=selitem.opening_stock*selitem.purchase_price  
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'items':items,
                    'selitem':selitem,
                    'stock_value':stock_value,
                    'latest_item_id':filtered_data,
                    'est_comments':est_comments
                    
            } 
                return render(request,'zohomodules/items/itemsoverview.html',content)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
         

    # Sort the items by the 'item_name' field
                items=Items.objects.filter(activation_tag='inactive',company=dash_details)  
                selitem=Items.objects.get(id=pk)
                est_comments=Items_comments.objects.filter(Items=pk)
                stock_value=selitem.opening_stock*selitem.purchase_price  
                latest_date = Item_Transaction_History.objects.filter(items_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                filtered_data = Item_Transaction_History.objects.filter(Date=latest_date, items_id=pk)
                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                content = {
                    'details': dash_details,
                    'allmodules': allmodules,
                    'items':items,
                    'selitem':selitem,
                    'stock_value':stock_value,
                    'latest_item_id':filtered_data,
                    'est_comments':est_comments
                    
            } 
                return render(request,'zohomodules/items/itemsoverview.html',content) 

    
    #--------------------------------------------------- TINTO VIEW ITEMS END-------------------------------------------


        #--------------------------------------------------- TINTO VIEW CHART OF ACCOUNTS START-------------------------------------------
def addchartofaccounts(request):                                                                #new by tinto mt
        if 'login_id' in request.session:
            login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=login_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                item=Items.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                units = Unit.objects.filter(company=dash_details.company)
                accounts=Chart_of_Accounts.objects.filter(company=dash_details.company)
                context = {
                     'details': dash_details,
        
                    'allmodules': allmodules,
         
                }
                return render(request,'zohomodules/chartofaccount/addchartofaccounts.html',context)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            item=Items.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            units = Unit.objects.filter(company=dash_details)
            accounts=Chart_of_Accounts.objects.filter(company=dash_details)
            context = {
                    'details': dash_details,
          
                    'allmodules': allmodules,
           
            }
    
            return render(request,'zohomodules/chartofaccount/addchartofaccounts.html',context)


def chartofaccounts(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccount/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc=Chart_of_Accounts.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccount/chartofaccounts.html',content)
  

def create_account(request):                                                                #new by tinto mt
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        if request.method=='POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            account=Chart_of_Accounts.objects.all()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.logindetails=log_user
            b.action="Created"
            b.Date=date.today()
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.account_number = request.POST.get("account_number2",None)
            a.account_description = request.POST['description']
            if a.account_type=="Other Current Assets":

                a.sub_account = request.POST.get("sub_account",None)
                a.parent_account = request.POST.get("parent_account",None)
               

            if a.account_type=="Cash":
             
                a.sub_account = request.POST.get("sub_account22",None)
                a.parent_account = request.POST.get("parent_account22",None)
               

            if a.account_type=="Fixed Assets":
            
                a.sub_account = request.POST.get("sub_account33",None)
                a.parent_account = request.POST.get("parent_account33",None)
               
            
            if a.account_type=="Stock":
               
                a.sub_account = request.POST.get("sub_account44",None)
                a.parent_account = request.POST.get("parent_account44",None)
             
            
            if a.account_type=="Other Current Liability":
             
                a.sub_account = request.POST.get("sub_account55",None)
                a.parent_account = request.POST.get("parent_account55",None)
               
            if a.account_type=="Long Term Liability":
            
                a.sub_account = request.POST.get("sub_account66",None)
                a.parent_account = request.POST.get("parent_account66",None)
              
            
            if a.account_type=="Other Liability":
              
                a.sub_account = request.POST.get("sub_account77",None)
                a.parent_account = request.POST.get("parent_account77",None)
              
            if a.account_type=="Equity":
            
                a.sub_account = request.POST.get("sub_account88",None)
                a.parent_account = request.POST.get("parent_account88",None)
            
            
            if a.account_type=="Income":
             
                a.sub_account = request.POST.get("sub_account99",None)
                a.parent_account = request.POST.get("parent_account99",None)
              
            
            if a.account_type=="Expense":
             
                a.sub_account = request.POST.get("sub_account100",None)
                a.parent_account = request.POST.get("parent_account100",None)
              
            if a.account_type=="Cost Of Goods Sold":
              
                a.sub_account = request.POST.get("sub_account111",None)
                a.parent_account = request.POST.get("parent_account111",None)
             
            if a.account_type=="Other Expense":
             
                a.sub_account = request.POST.get("sub_account222",None)
                a.parent_account = request.POST.get("parent_account222",None)
               
            account_type=request.POST.get("account_type",None)
            if account_type == 'Other Assets':
                a.description = 'Track special assets like goodwill and other intangible assets'
            if account_type == 'Other Current Assets':
                a.description = 'Any short term asset that can be converted into cash or cash equivalents easily Prepaid expenses Stocks and Mutual Funds'
            if account_type == 'Cash':
                a.description = 'To keep track of cash and other cash equivalents like petty cash, undeposited funds, etc., use an organized accounting system  financial software'
            if account_type == 'Bank':
                a.description = 'To keep track of bank accounts like Savings, Checking, and Money Market accounts.'
            if account_type == 'Fixed Asset':
                a.description = 'Any long-term investment or asset that cannot be easily converted into cash includes: Land and Buildings, Plant, Machinery, and Equipment, Computers, Furniture.'
            if account_type == 'Stock':
                a.description = 'To keep track of your inventory assets.'
            if account_type == 'Payment Clearing':
                a.description = 'To keep track of funds moving in and out via payment processors like Stripe, PayPal, etc.'
            if account_type == 'Other Liability':
                a.description = 'Obligation of an entity arising from past transactions or events which would require repayment.Tax to be paid Loan to be Repaid Accounts Payableetc.'
            if account_type == 'Other Current Liability':
                a.description = 'Any short term liability like: Customer Deposits Tax Payable'
            if account_type == 'Credit Card':
                a.description = 'Create a trail of all your credit card transactions by creating a credit card account.'
            if account_type == 'Long Term Liability':
                a.description = 'Liabilities that mature after a minimum period of one year like: Notes Payable Debentures Long Term Loans '
            if account_type == 'Overseas Tax Payable':
                a.description = 'Track your taxes in this account if your business sells digital services to foreign customers.'
            if account_type == 'Equity':
                a.description = 'Owners or stakeholders interest on the assets of the business after deducting all the liabilities.'
            if account_type == 'Income':
                a.description = 'Income or Revenue earned from normal business activities like sale of goods and services to customers.'
            if account_type == 'Other Income':
                a.description = 'Income or revenue earned from activities not directly related to your business like : Interest Earned Dividend Earned'
            if account_type == 'Expense':
                a.description = 'Reflects expenses incurred for running normal business operations, such as : Advertisements and Marketing Business Travel Expenses License Fees Utility Expenses'
            if account_type == 'Cost Of Goods Sold':
                a.description = 'This indicates the direct costs attributable to the production of the goods sold by a company such as: Material and Labor costs Cost of obtaining raw materials'
            if account_type == 'Other Expense':
                a.description = 'Track miscellaneous expenses incurred for activities other than primary business operations or create additional accounts to track default expenses like insurance or contribution towards charity.'
       

            
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name,company=c).exists():
                error='yes'
                messages.error(request,'Account with same name exsits !!!')
                return redirect('addchartofaccounts')
            else:
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                return redirect('chartofaccounts')
    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.account_number = request.POST.get("account_number2",None)
            a.account_description = request.POST['description']
            account_type=request.POST.get("account_type",None)
            if a.account_type=="Other Current Assets":

                a.sub_account = request.POST.get("sub_account",None)
                a.parent_account = request.POST.get("parent_account",None)
               

            if a.account_type=="Cash":
             
                a.sub_account = request.POST.get("sub_account22",None)
                a.parent_account = request.POST.get("parent_account22",None)
               

            if a.account_type=="Fixed Assets":
            
                a.sub_account = request.POST.get("sub_account33",None)
                a.parent_account = request.POST.get("parent_account33",None)
               
            
            if a.account_type=="Stock":
               
                a.sub_account = request.POST.get("sub_account44",None)
                a.parent_account = request.POST.get("parent_account44",None)
             
            
            if a.account_type=="Other Current Liability":
             
                a.sub_account = request.POST.get("sub_account55",None)
                a.parent_account = request.POST.get("parent_account55",None)
               
            if a.account_type=="Long Term Liability":
            
                a.sub_account = request.POST.get("sub_account66",None)
                a.parent_account = request.POST.get("parent_account66",None)
              
            
            if a.account_type=="Other Liability":
              
                a.sub_account = request.POST.get("sub_account77",None)
                a.parent_account = request.POST.get("parent_account77",None)
              
            if a.account_type=="Equity":
            
                a.sub_account = request.POST.get("sub_account88",None)
                a.parent_account = request.POST.get("parent_account88",None)
            
            
            if a.account_type=="Income":
             
                a.sub_account = request.POST.get("sub_account99",None)
                a.parent_account = request.POST.get("parent_account99",None)
              
            
            if a.account_type=="Expense":
             
                a.sub_account = request.POST.get("sub_account100",None)
                a.parent_account = request.POST.get("parent_account100",None)
              
            if a.account_type=="Cost Of Goods Sold":
              
                a.sub_account = request.POST.get("sub_account111",None)
                a.parent_account = request.POST.get("parent_account111",None)
             
            if a.account_type=="Other Expense":
             
                a.sub_account = request.POST.get("sub_account222",None)
                a.parent_account = request.POST.get("parent_account222",None)
               
            account_type=request.POST.get("account_type",None)
            if account_type == 'Other Assets':
                a.description = 'Track special assets like goodwill and other intangible assets'
            if account_type == 'Other Current Assets':
                a.description = 'Any short term asset that can be converted into cash or cash equivalents easily Prepaid expenses Stocks and Mutual Funds'
            if account_type == 'Cash':
                a.description = 'To keep track of cash and other cash equivalents like petty cash, undeposited funds, etc., use an organized accounting system  financial software'
            if account_type == 'Bank':
                a.description = 'To keep track of bank accounts like Savings, Checking, and Money Market accounts.'
            if account_type == 'Fixed Asset':
                a.description = 'Any long-term investment or asset that cannot be easily converted into cash includes: Land and Buildings, Plant, Machinery, and Equipment, Computers, Furniture.'
            if account_type == 'Stock':
                a.description = 'To keep track of your inventory assets.'
            if account_type == 'Payment Clearing':
                a.description = 'To keep track of funds moving in and out via payment processors like Stripe, PayPal, etc.'
            if account_type == 'Other Liability':
                a.description = 'Obligation of an entity arising from past transactions or events which would require repayment.Tax to be paid Loan to be Repaid Accounts Payableetc.'
            if account_type == 'Other Current Liability':
                a.description = 'Any short term liability like: Customer Deposits Tax Payable'
            if account_type == 'Credit Card':
                a.description = 'Create a trail of all your credit card transactions by creating a credit card account.'
            if account_type == 'Long Term Liability':
                a.description = 'Liabilities that mature after a minimum period of one year like: Notes Payable Debentures Long Term Loans '
            if account_type == 'Overseas Tax Payable':
                a.description = 'Track your taxes in this account if your business sells digital services to foreign customers.'
            if account_type == 'Equity':
                a.description = 'Owners or stakeholders interest on the assets of the business after deducting all the liabilities.'
            if account_type == 'Income':
                a.description = 'Income or Revenue earned from normal business activities like sale of goods and services to customers.'
            if account_type == 'Other Income':
                a.description = 'Income or revenue earned from activities not directly related to your business like : Interest Earned Dividend Earned'
            if account_type == 'Expense':
                a.description = 'Reflects expenses incurred for running normal business operations, such as : Advertisements and Marketing Business Travel Expenses License Fees Utility Expenses'
            if account_type == 'Cost Of Goods Sold':
                a.description = 'This indicates the direct costs attributable to the production of the goods sold by a company such as: Material and Labor costs Cost of obtaining raw materials'
            if account_type == 'Other Expense':
                a.description = 'Track miscellaneous expenses incurred for activities other than primary business operations or create additional accounts to track default expenses like insurance or contribution towards charity.'
       
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name,company=c).exists():
                error='yes'
                messages.error(request,'Account with same name exsits')
                return redirect('addchartofaccounts')
            else:
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                return redirect('chartofaccounts')

    return redirect('addchartofaccounts')

def chartofaccountsoverview(request,pk):                                                                #new by tinto mt
       if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=login_id)
        if log_details.user_type == 'Staff':
                    dash_details = StaffDetails.objects.get(login_details=log_details)

                    allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                
                    acc=Chart_of_Accounts.objects.filter(company=dash_details.company)  
                    selacc=Chart_of_Accounts.objects.get(id=pk)  
                    est_comments=chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)
                    latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
                    filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)
                    context = {
                        'details': dash_details,
                    
                        'allmodules': allmodules,
                        'acc':acc,
                        'selacc':selacc,
                        'latest_item_id':filtered_data,
                        'est_comments':est_comments,
                    }
                    return render(request, 'zohomodules/chartofaccount/chartofaccountsoverview.html',context)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
       
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            acc=Chart_of_Accounts.objects.filter(company=dash_details)  
            selacc=Chart_of_Accounts.objects.get(id=pk)  
            est_comments=chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)
            latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
            filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)
            context = {
                        'details': dash_details,
                    
                        'allmodules': allmodules,
                        'acc':acc,
                        'selacc':selacc,
                        'latest_item_id':filtered_data,
                        'est_comments':est_comments,
                    }
    
            return render(request, 'zohomodules/chartofaccount/chartofaccountsoverview.html',context)



   
        
    # acc=Chart_of_Accounts.objects.all()  
    # selacc=Chart_of_Accounts.objects.get(id=pk)  
    # est_comments=chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)
    # latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']    
    # filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)
    # return render(request, 'zohomodules/chartofaccounts/chartofaccountsoverview.html',{'acc':acc,'selacc':selacc,'latest_item_id':filtered_data,'est_comments':est_comments})


from django.shortcuts import render, redirect

def editchartofaccounts(request, pr):                                                                #new by tinto mt
    # Retrieve the chart of accounts entry
    

    # Check if 'company_id' is in the session
    if 'login_id' in request.session:
        login_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    
    # Retrieve the chart of accounts entry
    acc = get_object_or_404(Chart_of_Accounts, id=pr)

    # Check if 'company_id' is in the session

    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
     
        dash_details = CompanyDetails.objects.get(login_details=log_user)
       
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
   
        context = {
                    'acc': acc,
              
                    'details': dash_details,
                   
                    'allmodules': allmodules,
            }
       
    
        
        

        if request.method == 'POST':
        
            b=Chart_of_Accounts_History()
       
            b.company=dash_details
            b.logindetails=log_user
            b.action="Edited"
            b.Date=date.today()
            acc.login_details=log_user
            acc.company=dash_details
            # Update the chart of accounts entry with the form data
            acc.account_type = request.POST['account_type']
            acc.account_name = request.POST['account_name']
            acc.account_code = request.POST['account_code']
            acc.account_description = request.POST['description']
            
            # Save the changes
            acc.save()
            t=Chart_of_Accounts.objects.get(id=acc.id)
            b.chart_of_accounts=t
            b.save()

            # Redirect to another page after successful update
            return redirect('chartofaccountsoverview', pr)
        return render(request, 'zohomodules/chartofaccount/editchartofaccounts.html', context)
    if log_user.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_user)
                
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        
   
        context = {
                    'acc': acc,
              
                    'details': dash_details,
                   
                    'allmodules': allmodules,
            }
        if request.method=='POST':
         
            b=Chart_of_Accounts_History()
         
            c=dash_details.company
            b.company=c
            b.logindetails=log_user
            b.action="Edited"
            b.Date=date.today()
            acc.login_details=log_user
            acc.company=c
            # Update the chart of accounts entry with the form data
            acc.account_type = request.POST['account_type']
            acc.account_name = request.POST['account_name']
            acc.account_code = request.POST['account_code']
            acc.account_description = request.POST['description']
            
            # Save the changes
            acc.save()
            t=Chart_of_Accounts.objects.get(id=acc.id)
            b.chart_of_accounts=t
            b.save()

            # Redirect to another page after successful update
            return redirect('chartofaccountsoverview', pr)
        return render(request, 'zohomodules/chartofaccount/editchartofaccounts.html', context)

def deleteaccount(request,pl):                                                                #new by tinto mt
    acc=Chart_of_Accounts.objects.filter(id=pl)
    acc.delete()
    
    return redirect(chartofaccounts)


def acc_status_edit(request, pv):                                                                #new by tinto mt
    
    selacc = Chart_of_Accounts.objects.get(id=pv)

    if selacc.status == 'Active':
        selacc.status = 'inactive'
        selacc.save()
    elif selacc.status != 'Active':
        selacc.status = 'Active'
        selacc.save()

    selacc.save()

    return redirect('chartofaccountsoverview',pv)


def add_account_comment(request,pc):                                                                #new by tinto mt

    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        if request.method=="POST":
                    
                    com=chart_of_accounts_comments()
                    c = CompanyDetails.objects.get(login_details=company_id)
            
                    comment_comments=request.POST['comment']
                    com.company=c
                    com.logindetails=log_user
                    com.comments=comment_comments
                    acc=Chart_of_Accounts.objects.get(id=pc)
                    com.chart_of_accounts=acc
                    
                    com.save()
                    return redirect('chartofaccountsoverview',pc)

    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            com=chart_of_accounts_comments()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            
            comment_comments=request.POST['comment']
            com.company=c
            com.logindetails=log_user
            com.comments=comment_comments
            acc=Chart_of_Accounts.objects.get(id=pc)
            com.chart_of_accounts=acc
                    
            com.save()
            return redirect('chartofaccountsoverview',pc)


def delete_account_comment(request,ph,pr):                                                                #new by tinto mt
    acc=chart_of_accounts_comments.objects.filter(id=ph)
    acc.delete()
    ac=Chart_of_Accounts.objects.get(id=pr)
    
    return redirect(chartofaccountsoverview,ac.id)

def account_view_sort_by_name(request,pk):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc = Chart_of_Accounts.objects.filter(company=dash_details.company).order_by('account_name')
                selacc = Chart_of_Accounts.objects.get(id=pk)
                est_comments = chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)

                latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']
                filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc': acc, 
                        'selacc': selacc, 
                        'latest_item_id': filtered_data,
                        'est_comments': est_comments,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccount/chartofaccountsoverview.html',content)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
                acc = Chart_of_Accounts.objects.filter(company=dash_details).order_by('account_name')
                selacc = Chart_of_Accounts.objects.get(id=pk)
                est_comments = chart_of_accounts_comments.objects.filter(chart_of_accounts=pk)

                latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=pk).aggregate(latest_date=Max('Date'))['latest_date']
                filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=pk)
                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                content = {
                        'details': dash_details,
                        'acc': acc, 
                        'selacc': selacc, 
                        'latest_item_id': filtered_data,
                        'est_comments': est_comments,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccount/chartofaccountsoverview.html',content)


def shareaccountToEmail(request,pt):                                                                #new by tinto mt
    if request.user: 
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']
                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                print(emails_list)
                print('1')
           
           
                acc = Chart_of_Accounts.objects.get(id=pt)
                context = {
                
                    'selacc':acc,
                }
                print('2')
                template_path = 'zohomodules/chartofaccount/accountemailpdf.html'
                print('3')
                template = get_template(template_path)
                print('4')
                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)#, link_callback=fetch_resources)
                pdf = result.getvalue()
                print('5')
                filename = f'Account Details.pdf'
                subject = f"Account"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached Account Details. \n{email_message}\n\n--\nRegards,\n{acc.account_name}\n{acc.account_type}", from_email=settings.EMAIL_HOST_USER,to=emails_list)
                email.attach(filename,pdf,"application/pdf")
                email.send(fail_silently=False)
                msg = messages.success(request, 'Details has been shared via email successfully..!')
                return redirect(chartofaccountsoverview,pt)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(chartofaccountsoverview,pt)
        
        #--------------------------------------------------- TINTO VIEW CHART OF ACCOUNTS END-------------------------------------------
        
        
def chartofaccountsActive(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,status="active")
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccount/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc=Chart_of_Accounts.objects.filter(company=dash_details)
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccount/chartofaccounts.html',content)
            
def chartofaccountsInactive(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,status="inactive")
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccount/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc=Chart_of_Accounts.objects.filter(company=dash_details,status="inactive")
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccount/chartofaccounts.html',content)
            
            
#------------------------------------payroll employee--------------------------------
#------------------------------------------------GEORGE MATHEW---------------------------------------
def payroll_employee_create(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    blood=Bloodgroup.objects.all()
    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
    if log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
    content = {
            'details': dash_details,
            'allmodules': allmodules,
            'log_id':log_details,
            'blood':blood
            
    }
    return render(request,'zohomodules/payroll-employee/payroll_create_employee.html',content)
    
def employee_list(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        pay=payroll_employee.objects.filter(company=dash_details.company)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        content = {
                'details': dash_details,
                'pay':pay,
                'allmodules': allmodules,
                'log_id':log_details
        }
        return render(request,'zohomodules/payroll-employee/payroll_list.html',content)
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        pay=payroll_employee.objects.filter(company=dash_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        content = {
                'details': dash_details,
                'pay':pay,
                'allmodules': allmodules,
                'log_id':log_details
        }
        return render(request,'zohomodules/payroll-employee/payroll_list.html',content)
        
def employee_overview(request,pk):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type =='Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        pay=payroll_employee.objects.filter(company=dash_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        p=payroll_employee.objects.get(id=pk)
        comment_data=payroll_employee_comment.objects.filter(employee=pk)
        history=employee_history.objects.filter(employee=pk)
    if log_details.user_type =='Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        pay=payroll_employee.objects.filter(company=dash_details.company)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        p=payroll_employee.objects.get(id=pk)
        comment_data=payroll_employee_comment.objects.filter(employee=pk)
        history=employee_history.objects.filter(employee=pk)
    content = {
                'details': dash_details,
                'pay':pay,
                'p':p,
                'allmodules': allmodules,
                'comment':comment_data,
                'history':history,
                'log_id':log_details,
        }
    return render(request,'zohomodules/payroll-employee/overview_page.html',content)
    
def create_employee(request):
    if request.method=='POST':
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':    
            company_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
            title=request.POST['title']
            fname=request.POST['fname']
            lname=request.POST['lname']
            alias=request.POST['alias']
            joindate=request.POST['joindate']
            salarydate=request.POST['salary']
            saltype=request.POST['saltype']
            if (saltype == 'Fixed'):
                salary=request.POST['fsalary']
            else:
                salary=request.POST['vsalary']
            image=request.FILES.get('file')
            amountperhr=request.POST['amnthr']
            workhr=request.POST['hours'] 
            empnum=request.POST['empnum']
            if payroll_employee.objects.filter(emp_number=empnum,company=company_details):
                messages.info(request,'employee number all ready exists')
                return redirect('payroll_employee_create')
            designation = request.POST['designation']
            location=request.POST['location']
            gender=request.POST['gender']
            dob=request.POST['dob']
            blood=request.POST['blood']
            fmname=request.POST['fm_name']
            sname=request.POST['s_name']        
            add1=request.POST['address']
            add2=request.POST['address2']
            address=add1+" "+add2
            padd1=request.POST['paddress'] 
            padd2=request.POST['paddress2'] 
            paddress= padd1+padd2
            phone=request.POST['phone']
            ephone=request.POST['ephone']
            result_set1 = payroll_employee.objects.filter(company=company_details,Phone=phone)
            result_set2 = payroll_employee.objects.filter(company=company_details,emergency_phone=ephone)
            if result_set1:
                messages.error(request,'phone no already exists')
                return redirect('payroll_employee_create')
            if result_set2:
                messages.error(request,'phone no already exists')
                return redirect('payroll_employee_create')
            email=request.POST['email']
            result_set = payroll_employee.objects.filter(company=company_details,email=email)
            if result_set:
                messages.error(request,'email already exists')
                return redirect('payroll_employee_create')
            isdts=request.POST['tds']
            attach=request.FILES.get('attach')
            if isdts == '1':
                istdsval=request.POST['pora']
                if istdsval == 'Percentage':
                    tds=request.POST['pcnt']
                elif istdsval == 'Amount':
                    tds=request.POST['amnt']
            else:
                istdsval='No'
                tds = 0
            itn=request.POST['itn']
            an=request.POST['an']
            if payroll_employee.objects.filter(Aadhar=an,company=company_details):
                    messages.error(request,'Aadhra number already exists')
                    return redirect('payroll_employee_create')   
            uan=request.POST['uan'] 
            pfn=request.POST['pfn']
            pran=request.POST['pran']
            age=request.POST['age']
            bank=request.POST['bank']
            accno=request.POST['acc_no']       
            ifsc=request.POST['ifsc']       
            bname=request.POST['b_name']       
            branch=request.POST['branch']
            ttype=request.POST['ttype']
            if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
                payroll= payroll_employee(title=title,first_name=fname,last_name=lname,alias=alias,image=image,joindate=joindate,salary_type=saltype,salary=salary,age=age,
                            emp_number=empnum,designation=designation,location=location, gender=gender,dob=dob,blood=blood,parent=fmname,spouse_name=sname,workhr=workhr,
                            amountperhr = amountperhr, address=address,permanent_address=paddress ,Phone=phone,emergency_phone=ephone, email=email,Income_tax_no=itn,Aadhar=an,
                            UAN=uan,PFN=pfn,PRAN=pran,uploaded_file=attach,isTDS=istdsval,TDS_percentage=tds,salaryrange = salarydate,acc_no=accno,IFSC=ifsc,bank_name=bname,branch=branch,transaction_type=ttype,company=dash_details,login_details=log_details)
                payroll.save()
                history=employee_history(company=dash_details,login_details=log_details, employee=payroll,Action='CREATED')
                history.save()
                messages.info(request,'employee created')
                return redirect('employee_list')
        if log_details.user_type == 'Staff':
            company_details = StaffDetails.objects.get(login_details=log_details)
            title=request.POST['title']
            fname=request.POST['fname']
            lname=request.POST['lname']
            alias=request.POST['alias']
            joindate=request.POST['joindate']
            salarydate=request.POST['salary']
            saltype=request.POST['saltype']
            if (saltype == 'Fixed'):
                salary=request.POST['fsalary']
            else:
                salary=request.POST['vsalary']
            image=request.FILES.get('file')
            amountperhr=request.POST['amnthr']
            workhr=request.POST['hours'] 
            empnum=request.POST['empnum']
            if payroll_employee.objects.filter(emp_number=empnum,company=company_details.company):
                messages.info(request,'employee number all ready exists')
                return redirect('payroll_employee_create')
            designation = request.POST['designation']
            location=request.POST['location']
            gender=request.POST['gender']
            dob=request.POST['dob']
            blood=request.POST['blood']
            fmname=request.POST['fm_name']
            sname=request.POST['s_name']        
            add1=request.POST['address']
            add2=request.POST['address2']
            address=add1+" "+add2
            padd1=request.POST['paddress'] 
            padd2=request.POST['paddress2'] 
            paddress= padd1+padd2
            phone=request.POST['phone']
            ephone=request.POST['ephone']
            result_set1 = payroll_employee.objects.filter(company=company_details.company,Phone=phone)
            result_set2 = payroll_employee.objects.filter(company=company_details.company,emergency_phone=ephone)
            if result_set1:
                messages.error(request,'phone no already exists')
                return redirect('payroll_employee_create')
            if result_set2:
                messages.error(request,'emerency phone no already exists')
                return redirect('payroll_employee_create')
            email=request.POST['email']
            result_set = payroll_employee.objects.filter(company=company_details.company,email=email)
            if result_set:
                messages.error(request,'email already exists')
                return redirect('payroll_employee_create')
            isdts=request.POST['tds']
            attach=request.FILES.get('attach')
            if isdts == '1':
                istdsval=request.POST['pora']
                if istdsval == 'Percentage':
                    tds=request.POST['pcnt']
                elif istdsval == 'Amount':
                    tds=request.POST['amnt']
            else:
                istdsval='No'
                tds = 0
            itn=request.POST['itn']
            an=request.POST['an']
            if payroll_employee.objects.filter(Aadhar=an,company=company_details.company):
                    messages.error(request,'Aadhra number already exists')
                    return redirect('payroll_employee_create')   
            uan=request.POST['uan'] 
            pfn=request.POST['pfn']
            pran=request.POST['pran']
            age=request.POST['age']
            bank=request.POST['bank']
            accno=request.POST['acc_no']       
            ifsc=request.POST['ifsc']       
            bname=request.POST['b_name']       
            branch=request.POST['branch']
            ttype=request.POST['ttype']
            dash_details = StaffDetails.objects.get(login_details=log_details)
            payroll= payroll_employee(title=title,first_name=fname,last_name=lname,alias=alias,image=image,joindate=joindate,salary_type=saltype,salary=salary,age=age,
                         emp_number=empnum,designation=designation,location=location, gender=gender,dob=dob,blood=blood,parent=fmname,spouse_name=sname,workhr=workhr,
                         amountperhr = amountperhr, address=address,permanent_address=paddress ,Phone=phone,emergency_phone=ephone, email=email,Income_tax_no=itn,Aadhar=an,
                         UAN=uan,PFN=pfn,PRAN=pran,uploaded_file=attach,isTDS=istdsval,TDS_percentage=tds,salaryrange = salarydate,acc_no=accno,IFSC=ifsc,bank_name=bname,branch=branch,transaction_type=ttype,company=dash_details.company,login_details=log_details)
            payroll.save()
            history=employee_history(company=dash_details.company,login_details=log_details, employee=payroll,Action='CREATED')
            history.save()
            messages.info(request,'employee created')
            return redirect('employee_list')
    return redirect('payroll_employee_create')
    
def payroll_employee_edit(request,pk):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
    blood=Bloodgroup.objects.all()
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        p=payroll_employee.objects.get(id=pk)
    if log_details.user_type == 'Staff':
        dash_details = StaffDetails.objects.get(login_details=log_details)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        p=payroll_employee.objects.get(id=pk)
        
    print(p)
    content = {
            'details': dash_details,
            'allmodules': allmodules,
            'p':p,
            'log_id':log_details,
            'blood':blood
    }
    return render(request,'zohomodules/payroll-employee/edit_employee.html',content)
    
def do_payroll_edit(request,pk):
    if request.method=='POST':
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type =='Company':
            company_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)    
            title=request.POST['title']
            fname=request.POST['fname']
            lname=request.POST['lname']
            alias=request.POST['alias']
            joindate=request.POST['joindate']
            salarydate=request.POST['salary']
            saltype=request.POST['saltype']
            if (saltype == 'Fixed' or saltype =='Temporary'):
                salary=request.POST['fsalary']
            else:
                salary=request.POST['vsalary']
            image=request.FILES.get('file')
            amountperhr=request.POST['amnthr']
            workhr=request.POST['hours']
            empnum=request.POST['empnum']
            result_set2 = payroll_employee.objects.filter(company=company_details,emp_number=empnum).exclude(id=pk)
            if result_set2:
                messages.error(request,'employee number  already exists')
                return redirect('payroll_employee_edit',pk)
            designation = request.POST['designation']
            location=request.POST['location']
            gender=request.POST['gender']
            dob=request.POST['dob']
            blood=request.POST['blood']
            fmname=request.POST['fm_name']
            sname=request.POST['s_name']        
            add1=request.POST['address']
            add2=request.POST['address2']
            address=add1+" "+add2
            padd1=request.POST['paddress'] 
            padd2=request.POST['paddress2'] 
            paddress= padd1+padd2
            phone=request.POST['phone']
            ephone=request.POST['ephone']
            result_set1 = payroll_employee.objects.filter(company=company_details,Phone=phone).exclude(id=pk)
            result_set3 = payroll_employee.objects.filter(company=company_details,emergency_phone=phone).exclude(id=pk)
            if result_set1:
                messages.error(request,'phone no already exists')
                return redirect('payroll_employee_edit',pk)
            if result_set3:
                messages.error(request,'emergency phone no already exists')
                return redirect('payroll_employee_edit',pk)
            email=request.POST['email']
            result_set = payroll_employee.objects.filter(company=company_details,email=email).exclude(id=pk)
            if result_set:
                messages.error(request,'email already exists')
                return redirect('payroll_employee_edit',pk)
            isdts=request.POST['tds']
            attach=request.FILES.get('attach')
            if isdts == '1':
                istdsval=request.POST['pora']
                if istdsval == 'Percentage':
                    tds=request.POST['pcnt']
                elif istdsval == 'Amount':
                    tds=request.POST['amnt']
            else:
                istdsval='No'
                tds = 0
            itn=request.POST['itn']
            an=request.POST['an'] 
            if payroll_employee.objects.filter(Aadhar=an,company=company_details).exclude(id=pk):
                messages.error(request,'Aadhra number already exists')
                return redirect('payroll_employee_edit',pk)
            uan=request.POST['uan'] 
            pfn=request.POST['pfn']
            pran=request.POST['pran']
            age=request.POST['age']
            bank=request.POST['bank']
            accno=request.POST['acc_no']       
            ifsc=request.POST['ifsc']       
            bname=request.POST['b_name']       
            branch=request.POST['branch']
            ttype=request.POST['ttype']
            if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)
                payroll= payroll_employee.objects.get(id=pk)
                payroll.title=title
                payroll.first_name=fname
                payroll.last_name=lname
                payroll.alias=alias
                if len(request.FILES) != 0:
                    if image :
                        if payroll.image:
                            try:
                                # Check if the file exists before removing it
                                if os.path.exists(payroll.image.path):
                                    os.remove(payroll.image.path)
                            except Exception as e:
                                return redirect('payroll_employee_edit',pk)

                            # Assign the new file to payroll.image
                            payroll.image = image
                        else:
                            # Assign the new file to payroll.image
                            payroll.image = image
                payroll.joindate=joindate
                payroll.salary_type=saltype
                payroll.salary=salary
                age=age
                payroll.emp_number=empnum
                payroll.designation=designation
                payroll.location=location
                payroll.gender=gender
                payroll.dob=dob
                payroll.blood=blood
                payroll.parent=fmname
                payroll.spouse_name=sname
                payroll.workhr=workhr
                payroll.amountperhr = amountperhr
                payroll.address=address
                payroll.permanent_address=paddress
                payroll.Phone=phone
                payroll.emergency_phone=ephone
                payroll.email=email
                payroll.Income_tax_no=itn
                payroll.Aadhar=an
                payroll.UAN=uan
                payroll.PFN=pfn
                payroll.PRAN=pran
                if len(request.FILES) !=0:
                    if attach :
                        if payroll.uploaded_file:
                            try:
                                # Check if the file exists before removing it
                                if os.path.exists(payroll.uploaded_file.path):
                                    os.remove(payroll.uploaded_file.path)
                            except Exception as e:
                                return redirect('payroll_employee_edit',pk)

                            # Assign the new file to payroll.image
                            payroll.uploaded_file = attach
                        else:
                            # Assign the new file to payroll.image
                            payroll.uploaded_file = attach
                payroll.isTDS=istdsval
                payroll.TDS_percentage=tds
                payroll.salaryrange = salarydate
                payroll.acc_no=accno
                payroll.IFSC=ifsc
                payroll.bank_name=bname
                payroll.branch=branch
                payroll.transaction_type=ttype
                payroll.company=dash_details
                payroll.login_details=log_details
                payroll.save()
                history=employee_history(company=dash_details,login_details=log_details, employee=payroll,Action='EDITED')
                history.save()
                messages.info(request,'Updated')
                return redirect('employee_overview',pk)
        if log_details.user_type == 'Staff':
            if log_details.user_type =='Staff':
                company_details = StaffDetails.objects.get(login_details=log_details)    
                title=request.POST['title']
                fname=request.POST['fname']
                lname=request.POST['lname']
                alias=request.POST['alias']
                joindate=request.POST['joindate']
                salarydate=request.POST['salary']
                saltype=request.POST['saltype']
                if (saltype == 'Fixed' or saltype =='Temporary'):
                    salary=request.POST['fsalary']
                else:
                    salary=request.POST['vsalary']
                image=request.FILES.get('file')
                amountperhr=request.POST['amnthr']
                workhr=request.POST['hours']
                empnum=request.POST['empnum']
                result_set2 = payroll_employee.objects.filter(company=company_details.company,emp_number=empnum).exclude(id=pk)
                if result_set2:
                    messages.error(request,'employee number  already exists')
                    return redirect('payroll_employee_edit',pk)
                designation = request.POST['designation']
                location=request.POST['location']
                gender=request.POST['gender']
                dob=request.POST['dob']
                blood=request.POST['blood']
                fmname=request.POST['fm_name']
                sname=request.POST['s_name']        
                add1=request.POST['address']
                add2=request.POST['address2']
                address=add1+" "+add2
                padd1=request.POST['paddress'] 
                padd2=request.POST['paddress2'] 
                paddress= padd1+padd2
                phone=request.POST['phone']
                ephone=request.POST['ephone']
                result_set1 = payroll_employee.objects.filter(company=company_details.company,Phone=phone).exclude(id=pk)
                result_set3 = payroll_employee.objects.filter(company=company_details.company,emergency_phone=ephone).exclude(id=pk)
                if result_set1:
                    messages.error(request,'phone no already exists')
                    return redirect('payroll_employee_edit',pk)
                if result_set3:
                    messages.error(request,'emergency phone no already exists')
                    return redirect('payroll_employee_edit',pk)
                email=request.POST['email']
                result_set = payroll_employee.objects.filter(company=company_details.company,email=email).exclude(id=pk)
                if result_set:
                    messages.error(request,'email already exists')
                    return redirect('payroll_employee_edit',pk)
                isdts=request.POST['tds']
                attach=request.FILES.get('attach')
                if isdts == '1':
                    istdsval=request.POST['pora']
                    if istdsval == 'Percentage':
                        tds=request.POST['pcnt']
                    elif istdsval == 'Amount':
                        tds=request.POST['amnt']
                else:
                    istdsval='No'
                    tds = 0
                itn=request.POST['itn']
                an=request.POST['an'] 
                if payroll_employee.objects.filter(Aadhar=an,company=company_details.company).exclude(id=pk):
                    messages.error(request,'Aadhra number already exists')
                    return redirect('payroll_employee_edit',pk)
                uan=request.POST['uan'] 
                pfn=request.POST['pfn']
                pran=request.POST['pran']
                age=request.POST['age']
                bank=request.POST['bank']
                accno=request.POST['acc_no']       
                ifsc=request.POST['ifsc']       
                bname=request.POST['b_name']       
                branch=request.POST['branch']
                ttype=request.POST['ttype']
                dash_details = StaffDetails.objects.get(login_details=log_details)
                payroll= payroll_employee.objects.get(id=pk)
                payroll.title=title
                payroll.first_name=fname
                payroll.last_name=lname
                payroll.alias=alias
                if len(request.FILES) != 0:
                    if image :
                        if payroll.image:
                            try:
                                # Check if the file exists before removing it
                                if os.path.exists(payroll.image.path):
                                    os.remove(payroll.image.path)
                            except Exception as e:
                                return redirect('payroll_employee_edit',pk)

                            # Assign the new file to payroll.image
                            payroll.image = image
                        else:
                            # Assign the new file to payroll.image
                            payroll.image = image
                payroll.joindate=joindate
                payroll.salary_type=saltype
                payroll.salary=salary
                age=age
                payroll.emp_number=empnum
                payroll.designation=designation
                payroll.location=location
                payroll.gender=gender
                payroll.dob=dob
                payroll.blood=blood
                payroll.parent=fmname
                payroll.spouse_name=sname
                payroll.workhr=workhr
                payroll.amountperhr = amountperhr
                payroll.address=address
                payroll.permanent_address=paddress
                payroll.Phone=phone
                payroll.emergency_phone=ephone
                payroll.email=email
                payroll.Income_tax_no=itn
                payroll.Aadhar=an
                payroll.UAN=uan
                payroll.PFN=pfn
                payroll.PRAN=pran
                if len(request.FILES) !=0:
                    if attach :
                        if payroll.uploaded_file:
                            try:
                                # Check if the file exists before removing it
                                if os.path.exists(payroll.uploaded_file.path):
                                    os.remove(payroll.uploaded_file.path)
                            except Exception as e:
                                return redirect('payroll_employee_edit',pk)

                            # Assign the new file to payroll.image
                            payroll.uploaded_file = attach
                        else:
                            # Assign the new file to payroll.image
                            payroll.uploaded_file = attach
                payroll.isTDS=istdsval
                payroll.TDS_percentage=tds
                payroll.salaryrange = salarydate
                payroll.acc_no=accno
                payroll.IFSC=ifsc
                payroll.bank_name=bname
                payroll.branch=branch
                payroll.transaction_type=ttype
                payroll.company=dash_details.company
                payroll.login_details=log_details
                payroll.save()
                history=employee_history(company=dash_details.company,login_details=log_details, employee=payroll,Action='EDITED')
                history.save()
                messages.info(request,'Updated')
                return redirect('employee_overview',pk)
    return redirect('employee_overview',pk)
    
def add_comment(request,pk):
    if request.method =='POST':
        comment_data=request.POST['comments']
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        payroll= payroll_employee.objects.get(id=pk) 
        data=payroll_employee_comment(comment=comment_data,login_details=log_details,employee=payroll)
        data.save()
        return redirect('employee_overview',pk)
    return redirect('employee_overview',pk)
    
def delete_commet(request,pk,pi):
    data=payroll_employee_comment.objects.get(id=pk)
    data.delete()
    return redirect('employee_overview',pi)
    
def delete_employee(request,pk):
    data=payroll_employee.objects.get(id=pk)
    data.delete()
    return redirect('employee_list')
    
def employee_status(request,pk):
    data=payroll_employee.objects.get(id=pk)
    if data.status == 'Active':
        data.status ='Inactive'
    elif data.status == 'Inactive':
        data.status ='Active'
    data.save()
    return redirect('employee_overview',pk)
    
def add_blood(request):
    if request.method == 'POST':
        blood = request.POST.get('blood')
        print(blood)

        # Validate input
        if not blood:
            return JsonResponse({'message': 'Invalid or missing blood group'})

        # Use get_or_create for simplicity
        if Bloodgroup.objects.filter(Blood_group=blood):
            return JsonResponse({'message': 'Blood group already exists'})
        Bloodgroup.objects.create(Blood_group=blood)
        data=Bloodgroup.objects.all()
        return JsonResponse({'message': 'Blood group added','blood' : blood})
        
def import_payroll_excel(request):
    print(1)
    print('hello')
    if request.method == 'POST' :
        if 'login_id' in request.session:
            log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
            dash_details = StaffDetails.objects.get(login_details=log_details)
            if 'empfile' in request.FILES:
                excel_bill = request.FILES['empfile']
                excel_b = load_workbook(excel_bill)
                eb = excel_b['Sheet1']
                for row_number1 in range(2, eb.max_row + 1):
                    billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
                    payroll=payroll_employee(title=billsheet[0],first_name=billsheet[1],last_name=billsheet[2],alias=billsheet[3],joindate=datetime.date.fromisoformat(billsheet[4]),salary_type=billsheet[6],salary=billsheet[9],
                                emp_number=billsheet[10],designation=billsheet[11],location=billsheet[12], gender=billsheet[13],dob=datetime.date.fromisoformat(billsheet[14]),blood=billsheet[15],parent=billsheet[16],spouse_name=billsheet[17],workhr=billsheet[8],
                                amountperhr = billsheet[7], address=billsheet[19],permanent_address=billsheet[18],Phone=billsheet[20],emergency_phone=billsheet[21], email=billsheet[22],Income_tax_no=billsheet[32],Aadhar=billsheet[33],
                                UAN=billsheet[34],PFN=billsheet[35],PRAN=billsheet[36],isTDS=billsheet[29],TDS_percentage=billsheet[30],salaryrange = billsheet[5],acc_no=billsheet[24],IFSC=billsheet[25],bank_name=billsheet[26],branch=billsheet[27],transaction_type=billsheet[28],company=dash_details.company,login_details=log_details)
                    payroll.save()
                    history=employee_history(company=dash_details.company,login_details=log_details, employee=payroll,Action='IMPORTED')
                    history.save()
                    messages.warning(request,'file imported')
                    return redirect('employee_list')
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            if 'empfile' in request.FILES:
                excel_bill = request.FILES['empfile']
                excel_b = load_workbook(excel_bill)
                eb = excel_b['Sheet1']
                for row_number1 in range(2, eb.max_row + 1):
                    billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
                    payroll=payroll_employee(title=billsheet[0],first_name=billsheet[1],last_name=billsheet[2],alias=billsheet[3],joindate=billsheet[4],salary_type=billsheet[6],salary=billsheet[9],
                                emp_number=billsheet[10],designation=billsheet[11],location=billsheet[12], gender=billsheet[13],dob=billsheet[14],blood=billsheet[15],parent=billsheet[16],spouse_name=billsheet[17],workhr=billsheet[8],
                                amountperhr = billsheet[7], address=billsheet[19],permanent_address=billsheet[18],Phone=billsheet[20],emergency_phone=billsheet[21], email=billsheet[22],Income_tax_no=billsheet[32],Aadhar=billsheet[33],
                                UAN=billsheet[34],PFN=billsheet[35],PRAN=billsheet[36],isTDS=billsheet[29],TDS_percentage=billsheet[30],salaryrange = billsheet[5],acc_no=billsheet[24],IFSC=billsheet[25],bank_name=billsheet[26],branch=billsheet[27],transaction_type=billsheet[28],company=dash_details,login_details=log_details)
                    payroll.save()
                    history=employee_history(company=dash_details,login_details=log_details, employee=payroll,Action='IMPORTED')
                    history.save()
                    messages.warning(request,'file imported')
                    return redirect('employee_list')
    messages.error(request,'File upload Failed!11')
    return redirect('employee_list')
    
def add_file(request,pk):
    if request.method == 'POST':
        data=request.FILES.get('file')
        payroll=payroll_employee.objects.get(id=pk)
        if payroll.uploaded_file:
            try:
                                # Check if the file exists before removing it
                if os.path.exists(payroll.uploaded_file.path):
                    os.remove(payroll.uploaded_file.path)
            except Exception as e:
                messages.error(request,'file upload error')
                return redirect('employee_overview',pk)

                            # Assign the new file to payroll.image
            payroll.uploaded_file = data
            payroll.save()
            messages.info(request,'fil uploaded')
            return redirect('employee_overview',pk)
        else:
            payroll.uploaded_file = data
            payroll.save()
        messages.info(request,'fil uploaded')
        return redirect('employee_overview',pk)
        
def shareemail(request,pk):
    try:
            if request.method == 'POST':
                emails_string = request.POST['email']

    
                emails_list = [email.strip() for email in emails_string.split(',')]
                print(emails_list)
                p=payroll_employee.objects.get(id=pk)
                        
                context = {'p':p}
                template_path = 'zohomodules/payroll-employee/mailoverview.html'
                template = get_template(template_path)
                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
                pdf = result.getvalue()
                filename = f'{p.first_name}details - {p.id}.pdf'
                subject = f"{p.first_name}{p.last_name}  - {p.id}-details"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached employee details - File-{p.first_name}{p.last_name} .\n--\nRegards,\n", from_email=settings.EMAIL_HOST_USER, to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)
                messages.success(request, 'over view page has been shared via email successfully..!')
                return redirect('employee_overview',pk)
    except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect('employee_overview',pk)
#----------------------------------------------------------end----------------------------------------


def accounts_asset_filter(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,account_type__in=["Other Current Asset", "Fixed Asset","Other Asset"])
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details,account_type__in=["Other Current Asset", "Fixed Asset","Other Asset"])
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        

def accounts_liability_filter(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,account_type__in=["Other Current Liability", "Other Liability","Long Term Liability","Credit card","Overseas Tax Payable"])
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details,account_type__in=["Other Current Liability", "Other Liability","Long Term Liability","Credit card","Overseas Tax Payable"])
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)


def accounts_equity_filter(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,account_type__in=["Equity"])
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details,account_type__in=["Equity"])
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        

def accounts_income_filter(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,account_type__in=["Income","Other Income"])
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details,account_type__in=["Income","Other Income"])
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
def accounts_expense_filter(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,account_type__in=["Expense","Other Expense","Cost of Goods Sold"])
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details,account_type__in=["Expense","Other Expense","Cost of Goods Sold"])
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
            
            
def account_view_sort_by_namelist(request):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                # acc=Chart_of_Accounts.objects.filter(company=dash_details.company)
                acc = Chart_of_Accounts.objects.filter(company=dash_details.company).order_by('account_name')
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc':acc,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
        if log_details.user_type == 'Company':
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            # acc=Chart_of_Accounts.objects.filter(company=dash_details)
            acc = Chart_of_Accounts.objects.filter(company=dash_details).order_by('account_name')
            allmodules= ZohoModules.objects.get(company=dash_details,status='New')
            content = {
                    'details': dash_details,
                    'acc': acc,
                    'allmodules': allmodules,
            }   
            return render(request,'zohomodules/chartofaccounts/chartofaccounts.html',content)
            
            
def account_view_filterActive(request,ph):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,status="active")
                selacc = Chart_of_Accounts.objects.get(id=ph)
                est_comments = chart_of_accounts_comments.objects.filter(chart_of_accounts=ph)

                latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=ph).aggregate(latest_date=Max('Date'))['latest_date']
                filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=ph)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc': acc, 
                        'selacc': selacc, 
                        'latest_item_id': filtered_data,
                        'est_comments': est_comments,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccountsoverview.html',content)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details,status="active")
                selacc = Chart_of_Accounts.objects.get(id=ph)
                est_comments = chart_of_accounts_comments.objects.filter(chart_of_accounts=ph)

                latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=ph).aggregate(latest_date=Max('Date'))['latest_date']
                filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=ph)
                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                content = {
                        'details': dash_details,
                        'acc': acc, 
                        'selacc': selacc, 
                        'latest_item_id': filtered_data,
                        'est_comments': est_comments,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccountsoverview.html',content)
                
                
def account_view_filterinActive(request,ph):                                                                #new by tinto mt
     if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Staff':
                dash_details = StaffDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details.company,status="inactive")
                selacc = Chart_of_Accounts.objects.get(id=ph)
                est_comments = chart_of_accounts_comments.objects.filter(chart_of_accounts=ph)

                latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=ph).aggregate(latest_date=Max('Date'))['latest_date']
                filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=ph)
                allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
                content = {
                        'details': dash_details,
                        'acc': acc, 
                        'selacc': selacc, 
                        'latest_item_id': filtered_data,
                        'est_comments': est_comments,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccountsoverview.html',content)
        if log_details.user_type == 'Company':
                dash_details = CompanyDetails.objects.get(login_details=log_details)
                acc=Chart_of_Accounts.objects.filter(company=dash_details,status="inactive")
                selacc = Chart_of_Accounts.objects.get(id=ph)
                est_comments = chart_of_accounts_comments.objects.filter(chart_of_accounts=ph)

                latest_date = Chart_of_Accounts_History.objects.filter(chart_of_accounts_id=ph).aggregate(latest_date=Max('Date'))['latest_date']
                filtered_data = Chart_of_Accounts_History.objects.filter(Date=latest_date, chart_of_accounts_id=ph)
                allmodules= ZohoModules.objects.get(company=dash_details,status='New')
                content = {
                        'details': dash_details,
                        'acc': acc, 
                        'selacc': selacc, 
                        'latest_item_id': filtered_data,
                        'est_comments': est_comments,
                        'allmodules': allmodules,
                }
                return render(request,'zohomodules/chartofaccounts/chartofaccountsoverview.html',content)
                
                
#----------------- Banking -----------------------------#

def bank_list(request):
    log_id = request.session['login_id']
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        cmp = CompanyDetails.objects.get(login_details = log_details)
        dash_details = CompanyDetails.objects.get(login_details=log_details)
    else:
        cmp = StaffDetails.objects.get(login_details = log_details).company
        dash_details = StaffDetails.objects.get(login_details=log_details)
    bnk = Banking.objects.filter(company = cmp)
    allmodules= ZohoModules.objects.get(company = cmp)
    return render(request, 'bank_list.html',{'bnk':bnk, 'allmodules':allmodules, 'details':dash_details})

def load_bank_create(request):
    log_id = request.session['login_id']
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        cmp = CompanyDetails.objects.get(login_details = log_details)
        dash_details = CompanyDetails.objects.get(login_details=log_details)
    else:
        cmp = StaffDetails.objects.get(login_details = log_details).company
        dash_details = StaffDetails.objects.get(login_details=log_details)
    bnk = Banking.objects.filter(company = cmp)
    allmodules= ZohoModules.objects.get(company = cmp)
    tod = datetime.now().strftime('%Y-%m-%d')
    return render(request, 'bank_create.html',{'tod':tod, 'allmodules':allmodules, 'details':dash_details})

def bank_create(request):
    if request.method == 'POST':
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            cmp = CompanyDetails.objects.get(login_details = log_details)
        else:
            cmp = StaffDetails.objects.get(login_details = log_details).company

        date = request.POST.get('date')
        name = request.POST.get('name')
        opn_bal = request.POST.get('opn_bal')
        if opn_bal != '':
            opn_bal = float(opn_bal)
        else:
            opn_bal = 0
        bal_type=request.POST.get('bal_type')
        branch= request.POST.get('branch')
        ac_no= request.POST.get('ac_no')
        ifsc=request.POST.get('ifsc')

        if Banking.objects.filter(company = cmp, bnk_acno = ac_no).exists():
            messages.info(request, 'Bank with this account number already exsist !!')
            return redirect('bank_list')

        bank = Banking.objects.create(
            login_details = log_details,
            company = cmp,
            bnk_name=name,
            bnk_bal_type = bal_type,
            bnk_opnbal=opn_bal,
            bnk_bal=opn_bal,
            bnk_branch=branch,
            bnk_acno=ac_no,
            bnk_ifsc=ifsc,
            date=date
        )

        bank.save()

        BankTransaction.objects.create( login_details=log_details, company=cmp, banking=bank, trans_amount=opn_bal, trans_adj_date=date, 
                                       trans_type='Opening Balance', trans_adj_type='', trans_adj_amount = opn_bal)

        BankingHistory.objects.create( login_details=log_details, company=cmp, banking=bank, hist_adj_amount = 0, hist_adj_date=date, hist_action='Created')

        return redirect('bank_list')
    
def bank_view(request, id):
    log_id = request.session['login_id']
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        cmp = CompanyDetails.objects.get(login_details = log_details)
        dash_details = CompanyDetails.objects.get(login_details=log_details)
    else:
        cmp = StaffDetails.objects.get(login_details = log_details).company
        dash_details = StaffDetails.objects.get(login_details=log_details)
    allmodules= ZohoModules.objects.get(company = cmp)
    tod = datetime.now().strftime('%Y-%m-%d')
    bnk_list = Banking.objects.filter(company=cmp)
    bnk = Banking.objects.get(id=id, company=cmp)
    show_div = request.GET.get('Transaction', False)
    trans = BankTransaction.objects.filter(company=cmp, banking = bnk)
    hist_f = BankingHistory.objects.get(company = cmp, banking = bnk, hist_action = 'Created')
    hist_l = BankingHistory.objects.filter(company = cmp, banking = bnk).order_by('-id').first()
    context = {'cmp': cmp, 'bnk': bnk, 'bnk_list':bnk_list, 'trans':trans, 'tod':tod, 'id':id, 'show_div':show_div,
               'allmodules':allmodules, 'details':dash_details, 'hist_f':hist_f, 'hist_l':hist_l}
    return render(request, 'bank_view.html', context)

def bank_transaction_create(request, id):
    if request.method == 'POST':
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            cmp = CompanyDetails.objects.get(login_details = log_details)
        else:
            cmp = StaffDetails.objects.get(login_details = log_details).company

        date = request.POST.get('date')
        origin = request.POST.get('origin')
        dest = request.POST.get('dest')
        amount = float(request.POST.get('amount', 0.0))
        description = request.POST.get('description')
        ttype = request.POST.get('type')
        adjtype = request.POST.get('adjtype') 
        adjacname = request.POST.get('adjacname')
                    
        if ttype == 'Bank To Cash Transfer':
            origin_bnk = Banking.objects.get(id=origin)
            bal = float(origin_bnk.bnk_bal) - float(amount)
            thist =  BankTransaction.objects.create(
                login_details = log_details,
                company = cmp,
                banking = origin_bnk,
                trans_cur_amount = origin_bnk.bnk_opnbal,
                trans_amount = amount,
                trans_adj_amount = bal,
                trans_desc = description,
                trans_adj_date = date,
                trans_type = 'Bank to Cash',
                trans_adj_type = 'Balance Decrease'
            )
            BankTransactionHistory.objects.create(login_details=log_details, company=cmp, transaction=thist, hist_cur_amount=origin_bnk.bnk_bal, hist_amount=amount, 
                                                  hist_adj_amount = bal, hist_adj_date=date, hist_action='Created')
            origin_bnk.bnk_bal = bal
            origin_bnk.save()

        if ttype == 'Cash To Bank Transfer':
            dest_bnk = Banking.objects.get(id=dest)
            bal = float(dest_bnk.bnk_bal) + float(amount)
            thist = BankTransaction.objects.create(
                login_details = log_details,
                company = cmp,
                banking = dest_bnk,
                trans_cur_amount = dest_bnk.bnk_bal,
                trans_amount = amount,
                trans_adj_amount = bal,
                trans_desc = description,
                trans_adj_date = date,
                trans_type = 'Cash to Bank',
                trans_adj_type = 'Balance Increase'
            )
            BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=thist, hist_cur_amount=dest_bnk.bnk_bal, hist_amount=amount, 
                                                  hist_adj_amount = bal, hist_adj_date=date, hist_action='Created')
            dest_bnk.bnk_bal = bal
            dest_bnk.save()
            
        if ttype == 'Bank To Bank Transfer':
            from_bank = Banking.objects.get(id=origin)
            to_bank = Banking.objects.get(id=dest)
            bal = float(from_bank.bnk_bal) - float(amount)
            thist1 = BankTransaction.objects.create(
                login_details = log_details,
                company = cmp,
                banking = from_bank,
                trans_cur_amount = from_bank.bnk_bal,
                trans_amount = amount,
                trans_adj_amount = bal,
                trans_desc = description,
                trans_adj_date = date,
                trans_type = 'Bank to Bank',
                trans_adj_type = 'Balance Decrease'
            )
            BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=thist1, hist_cur_amount=from_bank.bnk_bal, hist_amount=amount, 
                                                  hist_adj_amount = bal, hist_adj_date=date, hist_action='Created')
            from_bank.bnk_bal = bal
            from_bank.save()

            bal = float(to_bank.bnk_bal) + float(amount)
            thist2 = BankTransaction.objects.create(
                login_details = log_details,
                company = cmp,
                banking = to_bank,
                trans_cur_amount = to_bank.bnk_bal,
                trans_amount = amount,
                trans_adj_amount = bal,
                trans_desc = description,
                trans_adj_date = date,
                trans_type = 'Bank to Bank',
                trans_adj_type = 'Balance Increase'
            )
            BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=thist2, hist_cur_amount=to_bank.bnk_bal, hist_amount=amount, 
                                                  hist_adj_amount = bal, hist_adj_date=date, hist_action='Created')
            to_bank.bnk_bal = bal
            to_bank.save()

            thist1.bank_to_bank_no = thist2.id
            thist2.bank_to_bank_no = thist1.id
            thist1.save()
            thist2.save()
                    
        if ttype == 'Adjust Bank Balance':
            adj_bank = Banking.objects.get(id=adjacname)
            if adjtype == 'Reduce Balance':
                adj_type = 'Balance Decrease'
                bal = float(adj_bank.bnk_bal) - float(amount)
            else:
                adj_type = 'Balance Increase'
                bal = float(adj_bank.bnk_bal) + float(amount)
            thist = BankTransaction.objects.create(
                login_details = log_details,
                company = cmp,
                banking = adj_bank,
                trans_cur_amount = adj_bank.bnk_bal,
                trans_amount = amount,
                trans_adj_amount = bal,
                trans_desc = description,
                trans_adj_date = date,
                trans_type = 'Bank Adjustment',
                trans_adj_type = adj_type
            )
            BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=thist, hist_cur_amount = adj_bank.bnk_bal, hist_amount = amount, 
                                                  hist_adj_amount = bal, hist_adj_date=date, hist_action='Created')
            adj_bank.bnk_bal = bal
            adj_bank.save()
        url = reverse('bank_view', kwargs={'id': id}) + '?Transaction=True'
        return redirect(url)
    
def load_trans_details(request):
    id = request.POST.get('id')
    trans = BankTransaction.objects.get(id=id)
    bnk_org_name = ''
    bnk_dest_name = ''
    if trans.trans_type == 'Bank to Bank':
        dest_trans = BankTransaction.objects.get(id=trans.bank_to_bank_no)
        bnk_org_name = trans.banking.id
        bnk_dest_name = dest_trans.banking.id
    else:
        bnk_org_name = trans.banking.id
        bnk_dest_name = trans.banking.id

    return JsonResponse({'message':'success', 
                         'ttype':trans.trans_type,
                         'origin':bnk_org_name,
                         'destination':bnk_dest_name,
                         'trans_id':trans.id,
                         'trans_adj_type':trans.trans_adj_type,
                         'trans_desc':trans.trans_desc,
                         'trans_adj_date':trans.trans_adj_date,
                         'trans_amount':trans.trans_amount})

def bank_transaction_edit(request):
    if request.method == 'POST':
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            cmp = CompanyDetails.objects.get(login_details = log_details)
        else:
            cmp = StaffDetails.objects.get(login_details = log_details).company

        id = request.POST.get('id')
        origin = request.POST.get('origin')
        dest = request.POST.get('dest')
        amount = float((request.POST.get('amount',0.0)))
        edit_date = request.POST.get('edit_date')
        desc = request.POST.get('desc')
        ttype = request.POST.get('type')
        adjtype = request.POST.get('adjtype') 
        adjacname = request.POST.get('adjacname')
                    
        # If it is Bank to Cash Transfer
        if ttype == 'Bank To Cash Transfer':
            htrans = BankTransaction.objects.get(id=id)
            origin_bnk = Banking.objects.get(id=origin)
            # If we did not change the Bank while editing
            if htrans.banking.bnk_acno == origin_bnk.bnk_acno:
                # Changing balance using edit amount
                bal = float(htrans.trans_cur_amount) - float(amount)
                htrans.trans_amount = amount
                htrans.trans_adj_amount = bal
                htrans.trans_desc = desc
                htrans.trans_adj_date = edit_date
                htrans.save()
                # Creating a transaction history for the edit
                BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=htrans, hist_cur_amount = htrans.trans_cur_amount, 
                                                        hist_amount = amount, hist_adj_amount = bal, hist_action='Updated')
                newbal = bal
                # Obtaning all transaction done after the edited transaction for that particular bank
                trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans.banking, id__gt=id)
                # Creating transaction history for all the transactions done after the edited transaction by calculating new balance
                for t in trans_list:
                    t.trans_cur_amount = newbal
                    nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                          hist_amount = t.trans_amount,  hist_action='Updated')
                    if t.trans_adj_type == 'Balance Increase':
                        newbal = float(t.trans_amount) + newbal
                    else:
                        newbal = newbal - float(t.trans_amount)  
                    nhist.hist_adj_amount = newbal
                    nhist.save()
                    t.trans_adj_amount = newbal
                    t.save()
                    
                # Changing final balance for the bank
                origin_bnk.bnk_bal = newbal
                origin_bnk.save()
            # If we changed the Bank while editing
            else:
                origin_bnk = Banking.objects.get(id=origin)
                htrans = BankTransaction.objects.get(id=id)
                # Deleting Transaction history of the transaction
                BankTransactionHistory.objects.filter(transaction = htrans).delete()

                # Obtaning all transaction done after the edited transaction for that particular bank
                trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans.banking, id__gt=id)
                newbal = htrans.trans_cur_amount
                # Creating transaction history for all the transactions done after the edited transaction by calculating new balance
                for t in trans_list:
                    t.trans_cur_amount = newbal
                    nhist = BankTransactionHistory.objects.create(login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                    hist_amount = t.trans_amount,  hist_action='Updated')
                    if t.trans_adj_type == 'Balance Increase':
                        newbal = float(t.trans_amount) + newbal
                    else:
                        newbal = newbal - float(t.trans_amount)  
                    nhist.hist_adj_amount = newbal
                    nhist.save()
                    t.trans_adj_amount = newbal
                    t.save()
                # Changing final balance for the bank
                htrans.banking.bnk_bal = newbal
                htrans.banking.save()

                bal = float(origin_bnk.bnk_bal) - float(amount)
                # Creating a transaction for the new bank
                newtrans = BankTransaction.objects.create(
                    login_details=log_details, 
                    company=cmp,
                    banking = origin_bnk,
                    trans_type = htrans.trans_type,
                    trans_adj_type = htrans.trans_adj_type,
                    trans_cur_amount = origin_bnk.bnk_bal,
                    trans_amount = amount,
                    trans_adj_amount = bal,
                    trans_desc = desc,
                    trans_adj_date = edit_date,
                )
                origin_bnk.bnk_bal = bal
                origin_bnk.save()
                # Creating a transaction history for the new bank
                BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans, hist_cur_amount=newtrans.trans_cur_amount, 
                                                          hist_amount=amount, hist_adj_amount=bal, hist_action='Created')
                # Deleting old bank transaction
                htrans.delete()

        # If it is Cash to Bank Transfer
        if ttype == 'Cash To Bank Transfer':
            htrans = BankTransaction.objects.get(id=id)
            dest_bnk = Banking.objects.get(id=dest)
            if htrans.banking.bnk_acno == dest_bnk.bnk_acno:
                bal = float(htrans.trans_cur_amount) + float(amount)
                htrans.trans_amount = amount
                htrans.trans_adj_amount = bal
                htrans.trans_desc = desc
                htrans.trans_adj_date = edit_date
                htrans.save()
                BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=htrans, hist_cur_amount = htrans.trans_cur_amount, 
                                                        hist_amount = amount, hist_adj_amount = bal, hist_action='Updated')
                newbal = bal
                trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans.banking, id__gt=id)
                for t in trans_list:
                    t.trans_cur_amount = newbal
                    nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                          hist_amount = t.trans_amount,  hist_action='Updated')
                    if t.trans_adj_type == 'Balance Increase':
                        newbal = float(t.trans_amount) + newbal
                    else:
                        newbal = newbal - float(t.trans_amount)  
                    nhist.hist_adj_amount = newbal
                    nhist.save()
                    t.trans_adj_amount = newbal
                    t.save()
                dest_bnk.bnk_bal = newbal
                dest_bnk.save()
            else:
                dest_bnk = Banking.objects.get(id=dest)
                htrans = BankTransaction.objects.get(id=id)
                BankTransactionHistory.objects.filter(transaction = htrans).delete()

                trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans.banking, id__gt=id)
                newbal = htrans.trans_cur_amount
                for t in trans_list:
                    t.trans_cur_amount = newbal
                    nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                          hist_amount = t.trans_amount,  hist_action='Updated')
                    if t.trans_adj_type == 'Balance Increase':
                        newbal = float(t.trans_amount) + newbal
                    else:
                        newbal = newbal - float(t.trans_amount)  
                    nhist.hist_adj_amount = newbal
                    nhist.save()
                    t.trans_adj_amount = newbal
                    t.save()
                htrans.banking.bnk_bal = newbal
                htrans.banking.save()

                bal = float(dest_bnk.bnk_bal) + float(amount)
                newtrans = BankTransaction.objects.create(
                    login_details=log_details, 
                    company=cmp,
                    banking = dest_bnk,
                    trans_type = htrans.trans_type,
                    trans_adj_type = htrans.trans_adj_type,
                    trans_cur_amount = dest_bnk.bnk_bal,
                    trans_amount = amount,
                    trans_adj_amount = bal,
                    trans_desc = desc,
                    trans_adj_date = edit_date,
                )
                dest_bnk.bnk_bal = bal
                dest_bnk.save()
                BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans, hist_cur_amount=newtrans.trans_cur_amount, 
                                                          hist_amount=amount, hist_adj_amount=bal, hist_action='Created')
                htrans.delete()

        # If it is Bank to Bank Transfer   
        if ttype == 'Bank To Bank Transfer':
            htrans_in = BankTransaction.objects.get(id=id)
            htrans_de = BankTransaction.objects.get(id=htrans_in.bank_to_bank_no)

            if htrans_in.trans_adj_type == 'Balance Decrease':
                red_bal = float(htrans_in.trans_cur_amount) - float(amount)
                add_bal = float(htrans_de.trans_cur_amount) + float(amount)
            else:
                red_bal = float(htrans_in.trans_cur_amount) + float(amount)
                add_bal = float(htrans_de.trans_cur_amount) - float(amount)

            origin_bnk = Banking.objects.get(id=origin)
            dest_bnk = Banking.objects.get(id=dest)

            # If Origin Bank is not changed
            if htrans_in.banking.bnk_acno == origin_bnk.bnk_acno:
                # If destination bank is not changed
                if htrans_in.banking.bnk_acno == dest_bnk.bnk_acno:
                    htrans_in.trans_amount = amount
                    htrans_in.trans_adj_amount = red_bal
                    htrans_in.trans_desc = desc
                    htrans_in.trans_adj_date = edit_date
                    htrans_in.save()
                    BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=htrans_in, hist_cur_amount = htrans_in.trans_cur_amount, 
                                                            hist_amount = amount, hist_adj_amount = red_bal, hist_action='Updated')
                    newbal = red_bal
                    trans_list = BankTransaction.objects.filter(company=cmp, banking=origin_bnk, id__gt=htrans_in.id)
                    for t in trans_list:
                        t.trans_cur_amount = newbal
                        nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                            hist_amount = t.trans_amount,  hist_action='Updated')
                        if t.trans_adj_type == 'Balance Increase':
                            newbal = float(t.trans_amount) + newbal
                        else:
                            newbal = newbal - float(t.trans_amount)  
                        nhist.hist_adj_amount = newbal
                        nhist.save()
                        t.trans_adj_amount = newbal
                        t.save()
                    origin_bnk.bnk_bal = newbal
                    origin_bnk.save()

                    htrans_de.trans_amount = amount
                    htrans_de.trans_adj_amount = add_bal
                    htrans_de.trans_desc = desc
                    htrans_de.trans_adj_date = edit_date
                    htrans_de.save()
                    BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=htrans_de, hist_cur_amount = htrans_de.trans_cur_amount, 
                                                            hist_amount = amount, hist_adj_amount = add_bal, hist_action='Updated')
                    newbal = add_bal
                    trans_list = BankTransaction.objects.filter(company=cmp, banking=dest_bnk, id__gt=htrans_de.id)
                    for t in trans_list:
                        t.trans_cur_amount = newbal
                        nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                            hist_amount = t.trans_amount,  hist_action='Updated')
                        if t.trans_adj_type == 'Balance Increase':
                            newbal = float(t.trans_amount) + newbal
                        else:
                            newbal = newbal - float(t.trans_amount)  
                        nhist.hist_adj_amount = newbal
                        nhist.save()
                        t.trans_adj_amount = newbal
                        t.save()
                    dest_bnk.bnk_bal = newbal
                    dest_bnk.save()
                #if destination bank is changed
                else:
                    htrans_in.trans_amount = amount
                    htrans_in.trans_adj_amount = red_bal
                    htrans_in.trans_desc = desc
                    htrans_in.trans_adj_date = edit_date
                    htrans_in.save()
                    BankTransactionHistory.objects.create(login_details=log_details, company=cmp, transaction=htrans_in, hist_cur_amount=htrans_in.trans_cur_amount, 
                                                            hist_amount = amount, hist_adj_amount = red_bal, hist_action='Updated')
                    newbal = red_bal
                    trans_list = BankTransaction.objects.filter(company=cmp, banking=origin_bnk, id__gt=htrans_in.id)
                    for t in trans_list:
                        t.trans_cur_amount = newbal
                        nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                            hist_amount = t.trans_amount,  hist_action='Updated')
                        if t.trans_adj_type == 'Balance Increase':
                            newbal = float(t.trans_amount) + newbal
                        else:
                            newbal = newbal - float(t.trans_amount)  
                        nhist.hist_adj_amount = newbal
                        nhist.save()
                        t.trans_adj_amount = newbal
                        t.save()
                    origin_bnk.bnk_bal = newbal
                    origin_bnk.save()


                    BankTransactionHistory.objects.filter(transaction = htrans_de).delete()
                    trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_de.banking, id__gt=htrans_de.id)
                    newbal = htrans_de.trans_cur_amount
                    for t in trans_list:
                        t.trans_cur_amount = newbal
                        nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                            hist_amount = t.trans_amount,  hist_action='Updated')
                        if t.trans_adj_type == 'Balance Increase':
                            newbal = float(t.trans_amount) + newbal
                        else:
                            newbal = newbal - float(t.trans_amount)  
                        nhist.hist_adj_amount = newbal
                        nhist.save()
                        t.trans_adj_amount = newbal
                        t.save()
                    htrans_de.banking.bnk_bal = newbal
                    htrans_de.banking.save()

                    if htrans_de.trans_adj_type == 'Balance Decrease':
                        bal = float(dest_bnk.bnk_bal) - float(amount)
                    else:
                        bal = float(dest_bnk.bnk_bal) + float(amount)

                    newtrans = BankTransaction.objects.create(
                        login_details=log_details, 
                        company=cmp,
                        banking = dest_bnk,
                        trans_type = htrans_de.trans_type,
                        trans_adj_type = htrans_de.trans_adj_type,
                        origin = origin_bnk.bnk_name,
                        origin_accno = origin_bnk.bnk_acno,
                        destination = dest_bnk.bnk_name,
                        destination_accno = dest_bnk.bnk_acno,
                        trans_cur_amount = dest_bnk.bnk_bal,
                        trans_amount = amount,
                        trans_adj_amount = bal,
                        trans_desc = desc,
                        trans_adj_date = edit_date,
                        bank_to_bank_no = htrans_in.id
                    )
                    dest_bnk.bnk_bal = bal
                    dest_bnk.save()
                    BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans, hist_cur_amount=newtrans.trans_cur_amount, 
                                                            hist_amount=amount, hist_adj_amount=bal, hist_action='Created')
                    htrans_de.delete()
            # if origin bank is changed 
            else:
                # if destination bank is not changed
                if htrans_in.banking.bnk_acno == dest_bnk.bnk_acno:
                    htrans_de.trans_amount = amount
                    htrans_de.trans_adj_amount = add_bal
                    htrans_de.trans_desc = desc
                    htrans_de.trans_adj_date = edit_date
                    htrans_de.save()
                    BankTransactionHistory.objects.create(login_details=log_details, company=cmp, transaction=htrans_de, hist_cur_amount=htrans_de.trans_cur_amount, 
                                                            hist_amount = amount, hist_adj_amount = add_bal, hist_action='Updated')
                    newbal = add_bal
                    trans_list = BankTransaction.objects.filter(company=cmp, banking=dest_bnk, id__gt=htrans_de.id)
                    for t in trans_list:
                        t.trans_cur_amount = newbal
                        nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                            hist_amount = t.trans_amount,  hist_action='Updated')
                        if t.trans_adj_type == 'Balance Increase':
                            newbal = float(t.trans_amount) + newbal
                        else:
                            newbal = newbal - float(t.trans_amount)  
                        nhist.hist_adj_amount = newbal
                        nhist.save()
                        t.trans_adj_amount = newbal
                        t.save()
                    dest_bnk.bnk_bal = newbal
                    dest_bnk.save()

                    BankTransactionHistory.objects.filter(transaction = htrans_in).delete()
                    trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_in.banking, id__gt=htrans_in.id)
                    newbal = htrans_in.trans_cur_amount
                    for t in trans_list:
                        t.trans_cur_amount = newbal
                        nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                            hist_amount = t.trans_amount,  hist_action='Updated')
                        if t.trans_adj_type == 'Balance Increase':
                            newbal = float(t.trans_amount) + newbal
                        else:
                            newbal = newbal - float(t.trans_amount)  
                        nhist.hist_adj_amount = newbal
                        nhist.save()
                        t.trans_adj_amount = newbal
                        t.save()
                    htrans_in.banking.bnk_bal = newbal
                    htrans_in.banking.save()

                    if htrans_in.trans_adj_type == 'Balance Decrease':
                        bal = float(origin_bnk.bnk_bal) - float(amount)
                    else:
                        bal = float(origin_bnk.bnk_bal) + float(amount)

                    newtrans = BankTransaction.objects.create(
                        login_details=log_details, 
                        company=cmp,
                        banking = origin_bnk,
                        trans_type = htrans_in.trans_type,
                        trans_adj_type = htrans_in.trans_adj_type,
                        trans_cur_amount = origin_bnk.bnk_bal,
                        trans_amount = amount,
                        trans_adj_amount = bal,
                        trans_desc = desc,
                        trans_adj_date = edit_date,
                        bank_to_bank_no = htrans_de.id
                    )
                    origin_bnk.bnk_bal = bal
                    origin_bnk.save()
                    BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans, hist_cur_amount=newtrans.trans_cur_amount, 
                                                            hist_amount=amount, hist_adj_amount=bal, hist_action='Created')
                    htrans_in.delete()
                # if destination bank is changed
                else:
                    # if new origin bank is not previous destination bank and new destiantion bank is not previous origin bank
                    if origin_bnk.bnk_acno != htrans_in.banking.bnk_acno and dest_bnk.bnk_acno != htrans_in.banking.bnk_acno:
                        BankTransactionHistory.objects.filter(transaction = htrans_in).delete()
                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_in.banking, id__gt=htrans_in.id)
                        newbal = htrans_in.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_in.banking.bnk_bal = newbal
                        htrans_in.banking.save()

                        if htrans_in.trans_adj_type == 'Balance Decrease':
                            bal = float(origin_bnk.bnk_bal) - float(amount)
                        else:
                            bal = float(origin_bnk.bnk_bal) + float(amount)

                        newtrans_or = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = origin_bnk,
                            trans_type = htrans_in.trans_type,
                            trans_adj_type = htrans_in.trans_adj_type,
                            trans_cur_amount = origin_bnk.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        origin_bnk.bnk_bal = bal
                        origin_bnk.save()
                        BankTransactionHistory.objects.create(login_details=log_details, company=cmp, transaction=newtrans_or, hist_amount=amount, 
                                                              hist_cur_amount=newtrans_or.trans_cur_amount, hist_adj_amount=bal, hist_action='Created')
                        htrans_in.delete()

                        BankTransactionHistory.objects.filter(transaction = htrans_de).delete()
                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_de.banking, id__gt=htrans_de.id)
                        newbal = htrans_de.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                            hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_de.banking.bnk_bal = newbal
                        htrans_de.banking.save()

                        if htrans_de.trans_adj_type == 'Balance Decrease':
                            bal = float(dest_bnk.bnk_bal) - float(amount)
                        else:
                            bal = float(dest_bnk.bnk_bal) + float(amount)

                        newtrans_tar = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = dest_bnk,
                            trans_type = htrans_de.trans_type,
                            trans_adj_type = htrans_de.trans_adj_type,
                            trans_cur_amount = dest_bnk.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        dest_bnk.bnk_bal = bal
                        dest_bnk.save()
                        BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans_tar, hist_cur_amount=newtrans_tar.trans_cur_amount, 
                                                                hist_amount=amount, hist_adj_amount=bal, hist_action='Created')
                        htrans_de.delete()

                        newtrans_or.bank_to_bank_no = newtrans_tar.id
                        newtrans_tar.bank_to_bank_no = newtrans_or.id
                        newtrans_or.save()
                        newtrans_tar.save()
                    
                    # if new origin bank is equal to previous destination bank and new destiantion bank is not previous origin bank
                    elif origin_bnk.bnk_acno == htrans_in.banking.bnk_acno and dest_bnk.bnk_acno != htrans_in.banking.bnk_acno:
                        BankTransactionHistory.objects.filter(transaction = htrans_in).delete()
                        BankTransactionHistory.objects.filter(transaction = htrans_de).delete()

                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_in.banking, id__gt=htrans_in.id)
                        newbal = htrans_in.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_in.banking.bnk_bal = newbal
                        htrans_in.banking.save()

                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_de.banking, id__gt=htrans_de.id)
                        newbal = htrans_de.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_de.banking.bnk_bal = newbal
                        htrans_de.banking.save()
                                                                        
                        if htrans_de.trans_adj_type == 'Balance Decrease':
                            bal = float(htrans_de.banking.bnk_bal) - float(amount)
                        else:
                            bal = float(htrans_de.banking.bnk_bal) + float(amount)

                        newtrans_or = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = origin_bnk,
                            trans_type = htrans_de.trans_type,
                            trans_adj_type = htrans_de.trans_adj_type,
                            trans_cur_amount = htrans_de.banking.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        origin_bnk.bnk_bal = bal
                        origin_bnk.save()
                        BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans_or, hist_cur_amount=newtrans_or.trans_cur_amount, 
                                                                hist_amount=amount, hist_adj_amount=bal, hist_action='Created')

                        if htrans_in.trans_adj_type == 'Balance Decrease':
                            bal = float(htrans_in.banking.bnk_bal) - float(amount)
                        else:
                            bal = float(htrans_in.banking.bnk_bal) + float(amount)

                        newtrans_tar = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = dest_bnk,
                            trans_type = htrans_in.trans_type,
                            trans_adj_type = htrans_in.trans_adj_type,
                            trans_cur_amount = htrans_in.banking.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        dest_bnk.bnk_bal = bal
                        dest_bnk.save()
                        BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans_tar, hist_cur_amount=newtrans_tar.trans_cur_amount, 
                                                                hist_amount=amount, hist_adj_amount=bal, hist_action='Created')

                        newtrans_or.bank_to_bank_no = newtrans_tar.id
                        newtrans_tar.bank_to_bank_no = newtrans_or.id
                        newtrans_or.save()
                        newtrans_tar.save()
                        htrans_in.delete()
                        htrans_de.delete()

                    # if new origin bank is not previous destination bank and new destiantion bank is equal to previous origin bank
                    elif origin_bnk.bnk_acno != htrans_in.banking.bnk_acno and dest_bnk.bnk_acno == htrans_in.banking.bnk_acno:
                        BankTransactionHistory.objects.filter(transaction = htrans_in).delete()
                        BankTransactionHistory.objects.filter(transaction = htrans_de).delete()
                    
                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_in.banking, id__gt=htrans_in.id)
                        newbal = htrans_in.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_in.banking.bnk_bal = newbal
                        htrans_in.banking.save()

                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_de.banking, id__gt=htrans_de.id)
                        newbal = htrans_de.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_de.banking.bnk_bal = newbal
                        htrans_de.banking.save()
                                                                        
                        if htrans_in.trans_adj_type == 'Balance Decrease':
                            bal = float(htrans_in.banking.bnk_bal) - float(amount)
                        else:
                            bal = float(htrans_in.banking.bnk_bal) + float(amount)

                        newtrans_or = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = origin_bnk,
                            trans_type = htrans_in.trans_type,
                            trans_adj_type = htrans_in.trans_adj_type,
                            trans_cur_amount = htrans_in.banking.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        origin_bnk.bnk_bal = bal
                        origin_bnk.save()
                        BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans_or, hist_cur_amount=newtrans_or.trans_cur_amount, 
                                                                hist_amount=amount, hist_adj_amount=bal, hist_action='Created')

                        if htrans_de.trans_adj_type == 'Balance Decrease':
                            bal = float(htrans_de.banking.bnk_bal) - float(amount)
                        else:
                            bal = float(htrans_de.banking.bnk_bal) + float(amount)

                        newtrans_tar = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = dest_bnk,
                            trans_type = htrans_de.trans_type,
                            trans_adj_type = htrans_de.trans_adj_type,
                            trans_cur_amount = htrans_de.banking.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        dest_bnk.bnk_bal = bal
                        dest_bnk.save()
                        BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans_tar, hist_cur_amount=newtrans_tar.trans_cur_amount, 
                                                                hist_amount=amount, hist_adj_amount=bal, hist_action='Created')

                        newtrans_or.bank_to_bank_no = newtrans_tar.id
                        newtrans_tar.bank_to_bank_no = newtrans_or.id
                        newtrans_or.save()
                        newtrans_tar.save()
                        htrans_in.delete()
                        htrans_de.delete()

                    # if new origin bank is equal to previous destination bank and new destiantion bank is equal to previous origin bank
                    else:
                        BankTransactionHistory.objects.filter(transaction = htrans_in).delete()
                        BankTransactionHistory.objects.filter(transaction = htrans_de).delete()
                    
                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_in.banking, id__gt=htrans_in.id)
                        newbal = htrans_in.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_in.banking.bnk_bal = newbal
                        htrans_in.banking.save()

                        trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans_de.banking, id__gt=htrans_de.id)
                        newbal = htrans_de.trans_cur_amount
                        for t in trans_list:
                            t.trans_cur_amount = newbal
                            nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                                hist_amount = t.trans_amount,  hist_action='Updated')
                            if t.trans_adj_type == 'Balance Increase':
                                newbal = float(t.trans_amount) + newbal
                            else:
                                newbal = newbal - float(t.trans_amount)  
                            nhist.hist_adj_amount = newbal
                            nhist.save()
                            t.trans_adj_amount = newbal
                            t.save()
                        htrans_de.banking.bnk_bal = newbal
                        htrans_de.banking.save()
                                           
                        if htrans_in.trans_adj_type == 'Balance Decrease':
                            bal = float(htrans_in.banking.bnk_bal) - float(amount)
                        else:
                            bal = float(htrans_in.banking.bnk_bal) + float(amount)

                        newtrans_or = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = htrans_in.banking,
                            trans_type = htrans_in.trans_type,
                            trans_adj_type = htrans_in.trans_adj_type,
                            trans_cur_amount = htrans_in.banking.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        origin_bnk.bnk_bal = bal
                        origin_bnk.save()
                        BankTransactionHistory.objects.create(login_details=log_details, company=cmp, transaction=newtrans_or, hist_amount=amount, 
                                                              hist_cur_amount=newtrans_or.trans_cur_amount, hist_adj_amount=bal, hist_action='Created')

                        if htrans_de.trans_adj_type == 'Balance Decrease':
                            bal = float(htrans_de.banking.bnk_bal) - float(amount)
                        else:
                            bal = float(htrans_de.banking.bnk_bal) + float(amount)

                        newtrans_tar = BankTransaction.objects.create(
                            login_details=log_details, 
                            company=cmp,
                            banking = htrans_de.banking,
                            trans_type = htrans_de.trans_type,
                            trans_adj_type = htrans_de.trans_adj_type,
                            trans_cur_amount = htrans_de.banking.bnk_bal,
                            trans_amount = amount,
                            trans_adj_amount = bal,
                            trans_desc = desc,
                            trans_adj_date = edit_date,
                        )
                        dest_bnk.bnk_bal = bal
                        dest_bnk.save()
                        BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans_tar, hist_amount=amount, 
                                                                hist_cur_amount=newtrans_or.trans_cur_amount, hist_adj_amount=bal, hist_action='Created')

                        newtrans_or.bank_to_bank_no = newtrans_tar.id
                        newtrans_tar.bank_to_bank_no = newtrans_or.id
                        newtrans_or.save()
                        newtrans_tar.save()
                        htrans_in.delete()
                        htrans_de.delete()

        # If it is Adjust Bank Balance           
        if ttype == 'Adjust Bank Balance':
            htrans = BankTransaction.objects.get(id=id)
            adj_bank = Banking.objects.get(id=adjacname)
            if adjtype == 'Reduce Balance':
                adj_type = 'Balance Decrease'
                bal = float(htrans.trans_cur_amount) - float(amount)
            else:
                adj_type = 'Balance Increase'
                bal = float(htrans.trans_cur_amount) + float(amount)
            if htrans.banking.bnk_acno == adj_bank.bnk_acno:
                htrans.trans_amount = amount
                htrans.trans_adj_amount = bal
                htrans.trans_desc = desc
                htrans.trans_adj_date = edit_date
                htrans.trans_adj_type = adj_type
                htrans.save()
                BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=htrans, hist_cur_amount = htrans.trans_cur_amount, 
                                                        hist_amount = amount, hist_adj_amount = bal, hist_action='Updated')
                newbal = bal
                trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans.banking, id__gt=id)
                for t in trans_list:
                    t.trans_cur_amount = newbal
                    nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                          hist_amount = t.trans_amount,  hist_action='Updated')
                    if t.trans_adj_type == 'Balance Increase':
                        newbal = float(t.trans_amount) + newbal
                    else:
                        newbal = newbal - float(t.trans_amount)  
                    nhist.hist_adj_amount = newbal
                    nhist.save()
                    t.trans_adj_amount = newbal
                    t.save()
                adj_bank.bnk_bal = newbal
                adj_bank.save()
            else:
                htrans = BankTransaction.objects.get(id=id)
                adj_bank = Banking.objects.get(id=adjacname)
                BankTransactionHistory.objects.filter(transaction = htrans).delete()
                trans_list = BankTransaction.objects.filter(company=cmp, banking=htrans.banking, id__gt=id)
                newbal = htrans.trans_cur_amount
                for t in trans_list:
                    t.trans_cur_amount = newbal
                    nhist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal, 
                                                          hist_amount = t.trans_amount,  hist_action='Updated')
                    if t.trans_adj_type == 'Balance Increase':
                        newbal = float(t.trans_amount) + newbal
                    else:
                        newbal = newbal - float(t.trans_amount)  
                    nhist.hist_adj_amount = newbal
                    nhist.save()
                    t.trans_adj_amount = newbal
                    t.save()
                htrans.banking.bnk_bal = newbal
                htrans.banking.save()

                bal = float(adj_bank.bnk_bal) - float(amount)
                newtrans = BankTransaction.objects.create(
                    login_details=log_details, 
                    company=cmp,
                    banking = adj_bank,
                    trans_type = 'Bank Adjustment',
                    trans_adj_type = htrans.trans_adj_type,
                    trans_cur_amount = adj_bank.bnk_bal,
                    trans_amount = amount,
                    trans_adj_amount = bal,
                    trans_desc = desc,
                    trans_adj_date = edit_date,
                )
                adj_bank.bnk_bal = bal
                adj_bank.save()
                BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=newtrans, hist_cur_amount=newtrans.trans_cur_amount, 
                                                          hist_amount=amount, hist_adj_amount=bal, hist_action='Created')
                htrans.delete()
        return JsonResponse({'message':'success'})
    
def load_bank_edit(request, id):
    log_id = request.session['login_id']
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        cmp = CompanyDetails.objects.get(login_details = log_details)
        dash_details = CompanyDetails.objects.get(login_details=log_details)
    else:
        cmp = StaffDetails.objects.get(login_details = log_details).company
        dash_details = StaffDetails.objects.get(login_details=log_details)
    bnk = Banking.objects.filter(company = cmp)
    allmodules= ZohoModules.objects.get(company = cmp)
    bnk = Banking.objects.get(id=id)
    return render(request, 'bank_edit.html', {'bnk':bnk, 'allmodules':allmodules, 'details':dash_details})

def bank_edit(request,id):
    if request.method == 'POST':
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            cmp = CompanyDetails.objects.get(login_details = log_details)
        else:
            cmp = StaffDetails.objects.get(login_details = log_details).company

        bnk = Banking.objects.get(id=id)
        bnk.bnk_name = request.POST.get('name')
        bnk.bnk_branch = request.POST.get('branch')
        bnk.bnk_acno = request.POST.get('ac_no')
        bnk.bnk_ifsc = request.POST.get('ifsc')
        bnk.bnk_bal_type = request.POST.get('bal_type')
        bnk.date = request.POST.get('date')
        newbal = request.POST.get('opn_bal')
        if newbal != '':
            newbal = float(newbal)
        else:
            newbal = 0
        bnk.bnk_opnbal = newbal
        bnk.save()
        BankingHistory.objects.create(login_details=log_details, company=cmp, banking=bnk, hist_adj_amount=request.POST.get('opn_bal', 0.0), hist_action='Updated')

        trans_list = BankTransaction.objects.filter(company=cmp, banking=bnk)
        for t in trans_list:
            hist = BankTransactionHistory.objects.create( login_details=log_details, company=cmp, transaction=t, hist_cur_amount = newbal,
                                                            hist_amount = t.trans_amount,  hist_action='Updated')
            if t.trans_type == 'Opening Balance':
                t.trans_amount = newbal
                t.trans_adj_amount = newbal
                t.trans_adj_date = bnk.date
            else:
                t.trans_cur_amount = newbal
                t.trans_adj_date = bnk.date
                if t.trans_adj_type == 'Balance Increase':
                    newbal = float(t.trans_amount) + newbal
                else:
                    newbal = newbal - float(t.trans_amount)  
                t.trans_adj_amount = newbal
            t.save()
            hist.hist_adj_amount = newbal
            hist.save()
        bnk.bnk_bal = newbal
        bnk.save()
        return redirect('bank_view',id)
    return redirect('bank_list')

def load_bank_history(request,id):
    log_id = request.session['login_id']
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        cmp = CompanyDetails.objects.get(login_details = log_details)
        dash_details = CompanyDetails.objects.get(login_details=log_details)
    else:
        cmp = StaffDetails.objects.get(login_details = log_details).company
        dash_details = StaffDetails.objects.get(login_details=log_details)
    bnk = Banking.objects.filter(company = cmp)
    allmodules= ZohoModules.objects.get(company = cmp)
    bhis = BankingHistory.objects.filter(company=cmp, banking=id)
    return render(request, 'bank_history.html', {'allmodules':allmodules, 'id':id, 'bhis':bhis, 'details':dash_details})

def load_bank_trans_history(request,id):
    log_id = request.session['login_id']
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        cmp = CompanyDetails.objects.get(login_details = log_details)
        dash_details = CompanyDetails.objects.get(login_details=log_details)
    else:
        cmp = StaffDetails.objects.get(login_details = log_details).company
        dash_details = StaffDetails.objects.get(login_details=log_details)
    bnk = Banking.objects.filter(company = cmp)
    allmodules= ZohoModules.objects.get(company = cmp)
    thist = BankTransactionHistory.objects.filter(transaction=id)
    bnk_id = thist[0].transaction.banking.id
    return render(request, 'bank_trans_history.html', {'allmodules':allmodules, 'thist':thist, 'id':bnk_id, 'details':dash_details})

def delete_banking(request, id):
    bnk = Banking.objects.get(id=id)
    if BankTransaction.objects.filter(banking=bnk).count() > 1:
        messages.info(request,'This bank cannot be deleted as it has done some transactions !!')
        return redirect('bank_view',id)
    bnk.delete()
    return redirect ('bank_list')

def delete_transaction(request,id):
    log_id = request.session['login_id']
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type == 'Company':
        cmp = CompanyDetails.objects.get(login_details = log_details)
    else:
        cmp = StaffDetails.objects.get(login_details = log_details).company
    trans = BankTransaction.objects.get(id=id)
    bnk_id = trans.banking.id

    if trans.trans_type == 'Bank to Bank':
        trans_list = BankTransaction.objects.filter(company=cmp, banking=trans.banking, id__gt=id)
        bal = trans.trans_cur_amount
        for t in trans_list:
            nhist = BankTransactionHistory.objects.create(login_details=log_details, company=cmp, transaction=t, hist_cur_amount = bal, 
                                                    hist_amount = t.trans_amount,  hist_action='Updated')
            if t.trans_adj_type == 'Balance Increase':
                newbal = float(t.trans_amount) + bal
            else:
                newbal = bal - float(t.trans_amount)  
            nhist.hist_adj_amount = newbal
            nhist.save()
            t.trans_cur_amount = bal
            t.trans_adj_amount = newbal
            t.save()
            bal = newbal
        trans.banking.bnk_bal = bal
        trans.banking.save()
        trans.delete()

        trans_nxt = BankTransaction.objects.get(id=trans.bank_to_bank_no)
        trans_list = BankTransaction.objects.filter(company=cmp, banking=trans_nxt.banking, id__gt=trans_nxt.id)
        bal = trans_nxt.trans_cur_amount
        for t in trans_list:
            nhist = BankTransactionHistory.objects.create(login_details=log_details, company=cmp, transaction=t, hist_cur_amount = bal, 
                                                    hist_amount = t.trans_amount,  hist_action='Updated')
            if t.trans_adj_type == 'Balance Increase':
                newbal = float(t.trans_amount) + bal
            else:
                newbal = bal - float(t.trans_amount)  
            nhist.hist_adj_amount = newbal
            nhist.save()
            t.trans_cur_amount = bal
            t.trans_adj_amount = newbal
            t.save()
            bal = newbal
        trans_nxt.banking.bnk_bal = bal
        trans_nxt.banking.save()
        trans_nxt.delete()

    else:
        trans_list = BankTransaction.objects.filter(company=cmp, banking=trans.banking, id__gt=id)
        bal = trans.trans_cur_amount
        for t in trans_list:
            nhist = BankTransactionHistory.objects.create(login_details=log_details, company=cmp, transaction=t, hist_cur_amount = bal, 
                                                    hist_amount = t.trans_amount,  hist_action='Updated')
            if t.trans_adj_type == 'Balance Increase':
                newbal = float(t.trans_amount) + bal
            else:
                newbal = bal - float(t.trans_amount)  
            nhist.hist_adj_amount = newbal
            nhist.save()
            t.trans_cur_amount = bal
            t.trans_adj_amount = newbal
            t.save()
            bal = newbal
        trans.banking.bnk_bal = bal
        trans.banking.save()
        trans.delete()
    messages.info(request,'Transaction Successfully Deleted !!')
    return redirect ('bank_view',bnk_id)

def banking_status(request,id):
    bnk = Banking.objects.get(id=id)
    if bnk.status == 'Active':
        bnk.status = 'Inactive'
    else:
        bnk.status = 'Active'
    bnk.save()
    return redirect('bank_view',id)

def bank_attachfile(request,id):
    if request.method == 'POST':
        bnk = Banking.objects.get(id=id) 
        bnk.document = request.POST.get('file')
        if len(request.FILES) != 0:
            bnk.document = request.FILES['file']
        bnk.save()
        return redirect('bank_view',id)
    
def send_bank_transaction(request,id):
    if request.method == 'POST':
        log_id = request.session['login_id']
        log_details= LoginDetails.objects.get(id=log_id)
        if log_details.user_type == 'Company':
            cmp = CompanyDetails.objects.get(login_details = log_details)
        else:
            dash_details = StaffDetails.objects.get(login_details=log_details)
        bnk = Banking.objects.get(id=id, company=cmp)
        trans = BankTransaction.objects.filter(company=cmp, banking=bnk)
        context = { 'bnk':bnk, 'trans':trans}

        emails_string = request.POST['mail']
        cemail = [email.strip() for email in emails_string.split(',')]
        template_path = 'bank_trans_template.html'
        template = get_template(template_path)
        html  = template.render(context)
        
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
        pdf = result.getvalue()
        filename = f'Transactions.pdf'
        subject = f"Bank Transactions"
        email = EmailMessage(subject, f"Hi,\nPlease find below the attached bank transaction on the bank {bnk.bnk_name} with account number {bnk.bnk_acno}.\n", 
                             from_email=settings.EMAIL_HOST_USER, to=cemail)
        email.attach(filename, pdf, "application/pdf")
        email.send(fail_silently=False)
        
        messages.success(request, 'Bill has been shared via email successfully..!')
        return redirect('bank_view', id)
    
def company_gsttype_change(request):
    if 'login_id' in request.session:
        log_id = request.session['login_id']
        if 'login_id' not in request.session:
            return redirect('/')

        log_details = LoginDetails.objects.get(id=log_id)
        dash_details = CompanyDetails.objects.get(login_details=log_details,superadmin_approval=1,Distributor_approval=1)

        if request.method == 'POST':
            # Get data from the form
            
            gstno = request.POST.get('gstno')
            gsttype = request.POST.get('gsttype')

            # Check if gsttype is one of the specified values
            if gsttype in ['unregistered Business', 'Overseas', 'Consumer']:
                dash_details.gst_no = None
            else:
                if gstno:
                    dash_details.gst_no = gstno
                else:
                    messages.error(request,'GST Number is not entered*')
                    return redirect('company_profile_editpage')


            dash_details.gst_type = gsttype

            dash_details.save()
            messages.success(request,'GST Type changed')
            return redirect('company_profile_editpage')
        else:
            return redirect('company_profile_editpage')
    else:
        return redirect('/') 
        
        
#------------------- PRICE LIST MODULE ------------
#  display all price lists
def all_price_lists(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type=="Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        sort_option = request.GET.get('sort', 'all')  
        filter_option = request.GET.get('filter', 'all')
        if sort_option == 'name':
            price_lists = price_lists.order_by('name')
        elif sort_option == 'type':
            price_lists = price_lists.order_by('type')

        if filter_option == 'active':
            price_lists = price_lists.filter(status='Active')
        elif filter_option == 'inactive':
            price_lists = price_lists.filter(status='Inactive')
        context={
            'log_id':log_id,
            'log_details':log_details,
            'details':dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'sort_option': sort_option,
            'filter_option': filter_option,
        }
        return render(request,'zohomodules/price_list/all_price_lists.html',context)
    
    if log_details.user_type=="Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details.company)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        sort_option = request.GET.get('sort', 'all')  
        filter_option = request.GET.get('filter', 'all')
        if sort_option == 'name':
            price_lists = price_lists.order_by('name')
        elif sort_option == 'type':
            price_lists = price_lists.order_by('type')

        if filter_option == 'active':
            price_lists = price_lists.filter(status='Active')
        elif filter_option == 'inactive':
            price_lists = price_lists.filter(status='Inactive')
        context={
            'log_id':log_id,
            'log_details':log_details,
            'details':dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'sort_option': sort_option,
            'filter_option': filter_option,
        }
        return render(request,'zohomodules/price_list/all_price_lists.html',context)



def import_price_list(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    
    log_details = LoginDetails.objects.get(id=log_id)

    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)

        if request.method == 'POST' and request.FILES.get('price_list_file') and request.FILES.get('items_file'):
            price_list_file = request.FILES['price_list_file']
            items_file = request.FILES['items_file']

            try:
                # Read PriceList Excel file(price_list_file)
                price_list_df = pd.read_excel(price_list_file)

                # Create PriceList and PriceListItem instances
                for index, row in price_list_df.iterrows():
                    # Check if a PriceList with the same name already exists
                    existing_price_list = PriceList.objects.filter(name=row['NAME'], company=dash_details).first()
                    if existing_price_list:
                        messages.error(request, f'Error importing data: PriceList with name "{row["NAME"]}" already exists.')
                        continue

                    new_price_list = PriceList.objects.create(
                        name=row['NAME'],
                        type=row['TYPE'],
                        item_rate_type=row['ITEM_RATE_TYPE'], 
                        description=row['DESCRIPTION'],
                        percentage_type=row['PERCENTAGE_TYPE'],
                        percentage_value=row['PERCENTAGE_VALUE'],
                        round_off=row['ROUNDING'],
                        currency=row['CURRENCY'],
                        company=dash_details, 
                        login_details=log_details,
                    )
                    PriceListTransactionHistory.objects.create(
                        company=dash_details,
                        login_details=log_details,
                        price_list=new_price_list,
                        action='Created',
                    )

                    # Read Items Excel file(items_file) for each PriceList
                    items_df = pd.read_excel(items_file)
                    for item_index, item_row in items_df.iterrows():
                        item = Items.objects.filter(item_name=item_row['ITEM_NAME'], company=dash_details, activation_tag='active').first()
                        if item:
                            standard_rate = item.selling_price if new_price_list.type == 'Sales' else item.purchase_price
                            custom_rate1 = item_row.get('SELLING_CUSTOM_RATE') if new_price_list.type == 'Sales' else item_row.get('PURCHASE_CUSTOM_RATE')
                            custom_rate = standard_rate if new_price_list.item_rate_type == 'Percentage' else custom_rate1
                            if custom_rate is None or math.isnan(custom_rate) or custom_rate == '':
                                custom_rate = Decimal(standard_rate)
                            
                            PriceListItem.objects.create(
                                company=dash_details,
                                login_details=log_details,
                                price_list=new_price_list,
                                item=item,
                                standard_rate=standard_rate,
                                custom_rate=custom_rate,
                            )
                        else:
                            messages.warning(request, f'Skipping item "{item_row["ITEM_NAME"]}" in PriceList "{row["NAME"]}": Item is not active.')

                messages.success(request, 'Price List data imported successfully.')
                return redirect('all_price_lists')

            except Exception as e:
                messages.error(request, f'Error importing data: {str(e)}')

    
    if log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)

        if request.method == 'POST' and request.FILES.get('price_list_file') and request.FILES.get('items_file'):
            price_list_file = request.FILES['price_list_file']
            items_file = request.FILES['items_file']

            try:
                # Read PriceList Excel file(price_list_file)
                price_list_df = pd.read_excel(price_list_file)

                # Create PriceList and PriceListItem instances
                for index, row in price_list_df.iterrows():
                    # Check if a PriceList with the same name already exists
                    existing_price_list = PriceList.objects.filter(name=row['NAME'], company=dash_details.company).first()
                    if existing_price_list:
                        messages.error(request, f'Error importing data: PriceList with name "{row["NAME"]}" already exists.')
                        continue

                    new_price_list = PriceList.objects.create(
                        name=row['NAME'],
                        type=row['TYPE'],
                        item_rate_type=row['ITEM_RATE_TYPE'], 
                        description=row['DESCRIPTION'],
                        percentage_type=row['PERCENTAGE_TYPE'],
                        percentage_value=row['PERCENTAGE_VALUE'],
                        round_off=row['ROUNDING'],
                        currency=row['CURRENCY'],
                        company=dash_details.company, 
                        login_details=log_details,
                    )
                    PriceListTransactionHistory.objects.create(
                        company=dash_details.company,
                        login_details=log_details,
                        price_list=new_price_list,
                        action='Created',
                    )

                    # Read Items Excel file(items_file) for each PriceList
                    items_df = pd.read_excel(items_file)
                    for item_index, item_row in items_df.iterrows():
                        item = Items.objects.filter(item_name=item_row['ITEM_NAME'], company=dash_details.company, activation_tag='active').first()
                        if item:
                            standard_rate = item.selling_price if new_price_list.type == 'Sales' else item.purchase_price
                            custom_rate1 = item_row.get('SELLING_CUSTOM_RATE') if new_price_list.type == 'Sales' else item_row.get('PURCHASE_CUSTOM_RATE')
                            custom_rate = standard_rate if new_price_list.item_rate_type == 'Percentage' else custom_rate1
                            if custom_rate is None or math.isnan(custom_rate) or custom_rate == '':
                                custom_rate = Decimal(standard_rate)
                            
                            PriceListItem.objects.create(
                                company=dash_details.company,
                                login_details=log_details,
                                price_list=new_price_list,
                                item=item,
                                standard_rate=standard_rate,
                                custom_rate=custom_rate,
                            )
                        else:
                            messages.warning(request, f'Skipping item "{item_row["ITEM_NAME"]}" in PriceList "{row["NAME"]}": Item is not active.')

                messages.success(request, 'Price List data imported successfully.')
                return redirect('all_price_lists')

            except Exception as e:
                messages.error(request, f'Error importing data: {str(e)}')

    else:
        return redirect('/')

    return redirect('all_price_lists')

def create_price_list(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    
    log_details = LoginDetails.objects.get(id=log_id)

    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        allmodules = ZohoModules.objects.get(company=dash_details, status='New')
        items = Items.objects.filter(company=dash_details,activation_tag='active')

        if request.method == 'POST':
            name = request.POST['name']
            if PriceList.objects.filter(name=name, company=dash_details).exists():
                messages.error(request, f"A Price List with the name '{name}' already exists.")
                return redirect('create_price_list')
            
            new_price_list = PriceList.objects.create(
                name=name,
                type=request.POST['type'],
                item_rate_type=request.POST['item_rate_type'],
                description=request.POST['description'],
                percentage_type=request.POST['percentage_type'],
                percentage_value=request.POST['percentage_value'],
                round_off=request.POST['round_off'],
                currency=request.POST['currency'],
                company=dash_details,
                login_details=log_details,
            )

            PriceListTransactionHistory.objects.create(
                company=dash_details,
                login_details=log_details,
                price_list=new_price_list,
                action='Created',
            )
            custom_rates = request.POST.getlist('custom_rate')
            for item, custom_rate in zip(items, custom_rates):
                custom_rate = custom_rate if custom_rate else (item.selling_price if new_price_list.type == 'Sales' else item.purchase_price)
                standard_rate = item.selling_price if new_price_list.type == 'Sales' else item.purchase_price
                PriceListItem.objects.create(
                    company=dash_details,
                    login_details=log_details,
                    price_list=new_price_list,
                    item=item,
                    standard_rate=standard_rate,
                    custom_rate=custom_rate,
                )
            return redirect('all_price_lists')
        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'items': items,
        }
        return render(request, 'zohomodules/price_list/create_price_list.html', context)

    if log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
        items = Items.objects.filter(company=dash_details.company,activation_tag='active')

        if request.method == 'POST':
            name = request.POST['name']
            if PriceList.objects.filter(name=name, company=dash_details.company).exists():
                messages.error(request, f"A Price List with the name '{name}' already exists.")
                return redirect('create_price_list')
            
            new_price_list = PriceList.objects.create(
                name=name,
                type=request.POST['type'],
                item_rate_type=request.POST['item_rate_type'],
                description=request.POST['description'],
                percentage_type=request.POST['percentage_type'],
                percentage_value=request.POST['percentage_value'],
                round_off=request.POST['round_off'],
                currency=request.POST['currency'],
                company=dash_details.company,
                login_details=log_details
            )
            
            PriceListTransactionHistory.objects.create(
                company=dash_details.company,
                login_details=log_details,
                price_list=new_price_list,
                action='Created',
            )
            
            custom_rates = request.POST.getlist('custom_rate')
            for item, custom_rate in zip(items, custom_rates):
                custom_rate = custom_rate if custom_rate else (item.selling_price if new_price_list.type == 'Sales' else item.purchase_price)
                standard_rate = item.selling_price if new_price_list.type == 'Sales' else item.purchase_price
                PriceListItem.objects.create(
                    company=dash_details.company,
                    login_details=log_details,
                    price_list=new_price_list,
                    item=item,
                    standard_rate=standard_rate,
                    custom_rate=custom_rate,
                )

            return redirect('all_price_lists')

        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'items': items,
        }
        return render(request, 'zohomodules/price_list/create_price_list.html', context)
 

def edit_price_list(request, price_list_id):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details = LoginDetails.objects.get(id=log_id)
    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details)
        allmodules = ZohoModules.objects.get(company=dash_details, status='New')
        price_list = get_object_or_404(PriceList, id=price_list_id)
        items = PriceListItem.objects.filter(price_list=price_list)

        
        if request.method == 'POST':
            price_list.name = request.POST['name']
            price_list.type = request.POST['type']
            price_list.item_rate_type = request.POST['item_rate_type']
            price_list.description = request.POST['description']
            price_list.percentage_type = request.POST['percentage_type']
            price_list.percentage_value = request.POST['percentage_value']
            price_list.round_off = request.POST['round_off']
            price_list.currency = request.POST['currency']
            price_list.save()
            
            PriceListTransactionHistory.objects.create(
                company=dash_details,
                login_details=log_details,
                price_list=price_list,
                action='Edited',
            )
            
            # edit PriceListItem
            custom_rate = request.POST.getlist('custom_rate')
            items = PriceListItem.objects.filter(price_list=price_list)
            for item, custom_rate in zip(items, custom_rate):
                standard_rate = item.item.selling_price if price_list.type == 'Sales' else item.item.purchase_price
                item.standard_rate = standard_rate
                item.custom_rate = custom_rate
                item.save()
            
            
            return redirect('price_list_details', price_list_id=price_list_id)
        context = {
            'log_details': log_details,
            'details': dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'price_list': price_list,
            'items': items,
        }
        return render(request, 'zohomodules/price_list/edit_price_list.html', context)
    elif log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details.company)
        allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
        price_list = get_object_or_404(PriceList, id=price_list_id)
        items = PriceListItem.objects.filter(price_list=price_list)
        
        if request.method == 'POST':
            price_list.name = request.POST['name']
            price_list.type = request.POST['type']
            price_list.item_rate_type = request.POST['item_rate_type']
            price_list.description = request.POST['description']
            price_list.percentage_type = request.POST['percentage_type']
            price_list.percentage_value = request.POST['percentage_value']
            price_list.round_off = request.POST['round_off']
            price_list.currency = request.POST['currency']
            price_list.save()
            PriceListTransactionHistory.objects.create(
                    company=dash_details.company,
                    login_details=log_details,
                    price_list=price_list,
                    action='Edited',
                )
            
            # edit PriceListItem
            custom_rate = request.POST.getlist('custom_rate')
            for item, custom_rate in zip(items, custom_rate):
                    standard_rate = item.item.selling_price if price_list.type == 'Sales' else item.item.purchase_price
                    item.standard_rate = standard_rate
                    item.custom_rate = custom_rate
                    item.save()
            
            return redirect('price_list_details', price_list_id=price_list_id)

        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'price_list': price_list,
            'items':items,
        }
        return render(request, 'zohomodules/price_list/edit_price_list.html', context)


# display details of selected price list
def price_list_details(request, price_list_id):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    
    if log_details.user_type=="Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details)
        price_list = get_object_or_404(PriceList, id=price_list_id)
        comments = PriceListComment.objects.filter(price_list=price_list)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        sort_option = request.GET.get('sort', 'all')  
        filter_option = request.GET.get('filter', 'all')
        if sort_option == 'name':
            price_lists = price_lists.order_by('name')
        elif sort_option == 'type':
            price_lists = price_lists.order_by('type')

        if filter_option == 'active':
            price_lists = price_lists.filter(status='Active')
        elif filter_option == 'inactive':
            price_lists = price_lists.filter(status='Inactive')
        transaction_history = PriceListTransactionHistory.objects.filter(price_list=price_list)
        items = PriceListItem.objects.filter(company=dash_details, price_list=price_list)
        latest_transaction = PriceListTransactionHistory.objects.filter(price_list=price_list)

        context={
            'log_id':log_id,
            'log_details':log_details,
            'details':dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'price_list': price_list,
            'comments': comments,
            'sort_option': sort_option,
            'filter_option': filter_option,
            'latest_transaction': latest_transaction,
            'transaction_history': transaction_history,
            'items':items,
        }
        return render(request,'zohomodules/price_list/price_list_details.html',context)
    
    if log_details.user_type=="Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details.company)
        price_list = get_object_or_404(PriceList, id=price_list_id)
        comments = PriceListComment.objects.filter(price_list=price_list)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        sort_option = request.GET.get('sort', 'all')  
        filter_option = request.GET.get('filter', 'all')
        if sort_option == 'name':
            price_lists = price_lists.order_by('name')
        elif sort_option == 'type':
            price_lists = price_lists.order_by('type')

        if filter_option == 'active':
            price_lists = price_lists.filter(status='Active')
        elif filter_option == 'inactive':
            price_lists = price_lists.filter(status='Inactive')
        transaction_history = PriceListTransactionHistory.objects.filter(price_list=price_list)
        items = PriceListItem.objects.filter(company=dash_details.company, price_list=price_list)
        context={
            'log_id':log_id,
            'log_details':log_details,
            'details':dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'comments': comments,
            'price_list': price_list,
            'sort_option': sort_option,
            'filter_option': filter_option,
            'transaction_history': transaction_history,
            'items':items,
        }
        return render(request,'zohomodules/price_list/price_list_details.html',context)
    

def delete_price_list(request, price_list_id):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type=="Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details)
        price_list = get_object_or_404(PriceList, id=price_list_id)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        price_list.delete()
        context={
            'details':dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'price_list': price_list,
        }
        return render(request,'zohomodules/price_list/all_price_lists.html',context)
    if log_details.user_type=="Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        price_lists = PriceList.objects.filter(company=dash_details.company)
        price_list = get_object_or_404(PriceList, id=price_list_id)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        price_list.delete()
        context={
            'details':dash_details,
            'allmodules': allmodules,
            'price_lists': price_lists,
            'price_list': price_list,
        }
        return render(request,'zohomodules/price_list/all_price_lists.html',context)


def toggle_price_list_status(request, price_list_id):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details = LoginDetails.objects.get(id=log_id)
    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        price_list = get_object_or_404(PriceList, id=price_list_id, company=dash_details)
        if price_list.status == 'Active':
            price_list.status = 'Inactive'
        else:
            price_list.status = 'Active'
        price_list.save()
        PriceListTransactionHistory.objects.create(
            company=dash_details,
            login_details=log_details,
            price_list=price_list,
            action='Edited' 
        )
        return redirect('price_list_details', price_list_id=price_list_id)
    if log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        price_list = get_object_or_404(PriceList, id=price_list_id, company=dash_details.company)
        if price_list.status == 'Active':
            price_list.status = 'Inactive'
        else:
            price_list.status = 'Active'
        price_list.save()
        PriceListTransactionHistory.objects.create(
            company=dash_details.company,
            login_details=log_details,
            price_list=price_list,
            action='Edited'  
        )
        return redirect('price_list_details', price_list_id=price_list_id)

def add_pricelist_comment(request, price_list_id):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details = LoginDetails.objects.get(id=log_id)
    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        price_list = get_object_or_404(PriceList, id=price_list_id, company=dash_details)
        if request.method == 'POST':
            comment = request.POST.get('comment_text')
            PriceListComment.objects.create(
                company=dash_details,
                login_details=log_details,
                price_list=price_list,
                comment=comment
            )
            
        return redirect('price_list_details', price_list_id=price_list_id)
    if log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        price_list = get_object_or_404(PriceList, id=price_list_id, company=dash_details.company)
        if request.method == 'POST':
            comment = request.POST.get('comment_text')
            PriceListComment.objects.create(
                company=dash_details.company,
                login_details=log_details,
                price_list=price_list,
                comment=comment
            )
        return redirect('price_list_details', price_list_id=price_list_id)

def delete_pricelist_comment(request, comment_id, price_list_id):
    comment = get_object_or_404(PriceListComment, id=comment_id)
    comment.delete()
    return redirect('price_list_details', price_list_id=price_list_id)





def whatsapp_pricelist(request, price_list_id):
    try:
        price_list = PriceList.objects.get(id=price_list_id)
        price_list_items = PriceListItem.objects.filter(price_list=price_list)

        context = {
            'price_list': price_list,
            'price_list_items': price_list_items,
        }

        # Render the template
        html = render(request, 'zohomodules/price_list/pdf_price_list.html', context).content

        # Create a PDF file
        pdf_file = BytesIO()
        pisa.pisaDocument(BytesIO(html), pdf_file)

        # Check if PDF generation was successful
        if pdf_file.tell():
            pdf_file.seek(0)

            # Save the PDF to the server's media folder
            pdf_filename = f"{price_list.name}_price_list.pdf"
            pdf_path = os.path.join('media', pdf_filename)
            with open(pdf_path, 'wb') as pdf_file_on_server:
                pdf_file_on_server.write(pdf_file.read())

            # Create a direct link to the saved PDF
            pdf_link = f"{request.scheme}://{request.get_host()}/{pdf_path}"

            # Create the WhatsApp message with a link to download the PDF
            whatsapp_message = f"Check out this price list: [Download PDF]{pdf_link}"

            # Create the WhatsApp link
            whatsapp_link = f"https://wa.me/?text={whatsapp_message}"

            # Return the WhatsApp link
            return redirect(whatsapp_link)

    except Exception as e:
        print(e)
        messages.error(request, f'{e}')

    # If there is an error or PDF generation fails, redirect to 'all_price_lists'
    return redirect('all_price_lists')



# email pricelist details(overview)
def email_pricelist(request, price_list_id):
    try:
        price_list = PriceList.objects.get(id=price_list_id)
        price_list_item = PriceListItem.objects.filter( price_list=price_list)

        if request.method == 'POST':
            emails_string = request.POST['email_ids']
            emails_list = [email.strip() for email in emails_string.split(',')]
            email_message = request.POST['email_message']

            context = {
                'price_list': price_list,
                'price_list_item': price_list_item,
            }

            template_path = 'zohomodules/price_list/pdf_price_list.html'
            template = get_template(template_path)
            html = template.render(context)
            result = BytesIO()
            pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
            pdf = result.getvalue()

            filename = f'Price_List_Details.pdf'
            subject = f"Price List Details: {price_list.name}"
            email = EmailMessage(subject, f"Hi,\nPlease find the attached Price List Details. \n{email_message}\n\n--\nRegards,\n{price_list.name}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
            email.attach(filename, pdf, "application/pdf")
            email.send(fail_silently=False)

            msg = messages.success(request, 'Details have been shared via email successfully..!')
            return redirect('price_list_details', price_list_id=price_list_id)  

    except Exception as e:
        print(e)
        messages.error(request, f'{e}')
        return redirect('all_price_lists')  

# dwnld pdf
def price_list_pdf(request, price_list_id):
    try:
        price_list = PriceList.objects.get(id=price_list_id)
        price_list_item = PriceListItem.objects.filter(price_list=price_list)

        context = {
            'price_list': price_list,
            'price_list_item': price_list_item,
        }

        template_path = 'zohomodules/price_list/pdf_price_list.html'
        template = get_template(template_path)
        html = template.render(context)
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
        pdf = result.getvalue()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{price_list.name}_Details.pdf"'
        response.write(pdf)
        return response
    except Exception as e:
        print(e)
        messages.error(request, f'{e}')
        return redirect('all_price_lists')

# upload attachment
def attach_file(request, price_list_id):
    price_list = PriceList.objects.get(pk=price_list_id)
    if request.method == 'POST':
        attachment = request.FILES.get('attachment')
        price_list.attachment = attachment
        price_list.save()
        return redirect('price_list_details', price_list_id=price_list.id)
    return HttpResponse("Invalid request method.")


#----------------------------------------------------------akshay--end--------------------------------------------------------


#===================================MANUAL JOURNAL==============================================

def manual_journal(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    print("hi")
    print(log_details)
    if log_details.user_type=="Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        print('hi')
        print(dash_details)
        journal = Journal.objects.filter(company=dash_details)
        print('cat')
        #print(journal)
        journal_entries = JournalEntry.objects.filter(journal__in=journal,company=dash_details)
        print(journal_entries)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        print('dog')
        print(allmodules)
        sort_option = request.GET.get('sort', 'all')  
        filter_option = request.GET.get('filter', 'all')
        if sort_option == 'journal_no':
            journal = journal.order_by('journal_no')
        elif sort_option == 'total_debit':
            journal = journal.order_by('total_debit')

        if filter_option == 'save':
            journal = journal.filter(status='save')
        elif filter_option == 'draft':
            journal = journal.filter(status='draft')
        context={
            'log_id':log_id,
            'log_details':log_details,
            'details':dash_details,
            'allmodules': allmodules,
            'journal': journal,
            'journal_entries':journal_entries,
            'sort_option': sort_option,
            'filter_option': filter_option,
        }
        
        return render(request,'zohomodules/manual_journal/manual_journal.html',context)
    
    if log_details.user_type=="Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        print('lilly')
        print(dash_details)
        journal = Journal.objects.filter(staff=dash_details)
        print('lotus')
        #print(journal)
        journal_entries = JournalEntry.objects.filter(journal__in=journal,staff=dash_details)
        print(journal_entries)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        print('rose')
        print(allmodules)
        sort_option = request.GET.get('sort', 'all')  
        filter_option = request.GET.get('filter', 'all')
        if sort_option == 'journal_nojournal_no':
            journal = journal.order_by('journal_no')
        elif sort_option == 'total_debit':
            journal = journal.order_by('total_debit')

        if filter_option == 'save':
            journal = journal.filter(status='save')
        elif filter_option == 'draft':
            journal = journal.filter(status='draft')
        context={
            'log_id':log_id,
            'log_details':log_details,
            'details':dash_details,
            'allmodules': allmodules,
            'journal': journal,
            'journal_entries':journal_entries,
            'sort_option': sort_option,
            'filter_option': filter_option,
        }
        
        return render(request,'zohomodules/manual_journal/manual_journal.html',context)

def import_journal_list(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    
    log_details = LoginDetails.objects.get(id=log_id)

    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)

        if request.method == 'POST' and request.FILES.get('journalfile') and request.FILES.get('accountfile'):
            journalfile = request.FILES['journalfile']
            accountfile = request.FILES['accountfile']

            try:
                # Read Journal Excel file(journalfile)
                journal_df = pd.read_excel(journalfile)

                # Create PriceList and PriceListItem instances
                for index, row in journal_df.iterrows():

                    new_journal = Journal.objects.create(
                        journal_no=row['JOURNAL_NO'],
                        reference_no=row['REFERENCE_NO'],
                        notes=row['NOTES'], 
                        currency=row['CURRENCY'],
                        journal_type=row['JOURNAL_TYPE'],
                        total_debit=row['TOTAL_DEBIT'],
                        total_credit=row['TOTAL_CREDIT'],
                        debit_difference=row['DEBIT_DIFFERENCE'],
                        credit_difference=row['CREDIT_DIFFERENCE'],
                        company=dash_details, 
                        login_details=log_details,
                    )
                    
                    JournalTransactionHistory.objects.create(
                        company=dash_details,
                        login_details=log_details,
                        journal=new_journal,
                        action='Created',
                    )

                    # Read account Excel file(account_file) for each Journal
                    account_df = pd.read_excel(accountfile)
                    for account_index, account_row in account_df.iterrows():     
                        JournalEntry.objects.create(
                            company=dash_details,
                            journal=new_journal,
                            account=account_row['ACCOUNT'],
                            description=account_row['DESCRIPTION'],
                            contact=account_row['CONTACT'],
                            debits=account_row['DEBITS'],
                            credits=account_row['CREDITS'],
                        )

                messages.success(request, 'Journal data imported successfully.')
                return redirect('manual_journal')

            except Exception as e:
                messages.error(request, f'Error importing data: {str(e)}')

    
    if log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)

        if request.method == 'POST' and request.FILES.get('journalfile') and request.FILES.get('accountfile'):
            journalfile = request.FILES['journalfile']
            accountfile = request.FILES['accountfile']

            try:
                # Read Journal Excel file(journalfile)
                journal_df = pd.read_excel(journalfile)

                # Create PriceList and PriceListItem instances
                for index, row in journal_df.iterrows():

                    new_journal = Journal.objects.create(
                        journal_no=row['JOURNAL_NO'],
                        reference_no=row['REFERENCE_NO'],
                        notes=row['NOTES'], 
                        currency=row['CURRENCY'],
                        journal_type=row['JOURNAL_TYPE'],
                        total_debit=row['TOTAL_DEBIT'],
                        total_credit=row['TOTAL_CREDIT'],
                        debit_difference=row['DEBIT_DIFFERENCE'],
                        credit_difference=row['CREDIT_DIFFERENCE'],
                        staff=dash_details, 
                        login_details=log_details,
                    )
                    
                    JournalTransactionHistory.objects.create(
                        staff=dash_details,
                        login_details=log_details,
                        journal=new_journal,
                        action='Created',
                    )

                    # Read account Excel file(account_file) for each Journal
                    account_df = pd.read_excel(accountfile)
                    for account_index, account_row in account_df.iterrows():     
                        JournalEntry.objects.create(
                            staff=dash_details,
                            journal=new_journal,
                            account=account_row['ACCOUNT'],
                            description=account_row['DESCRIPTION'],
                            contact=account_row['CONTACT'],
                            debits=account_row['DEBITS'],
                            credits=account_row['CREDITS'],
                        )

                messages.success(request, 'Journal data imported successfully.')
                return redirect('manual_journal')

            except Exception as e:
                messages.error(request, f'Error importing data: {str(e)}')

    else:
        return redirect('manual_journal')

    return redirect('manual_journal')


































def check_journal_num_valid(request):
    journals = JournalRecievedIdModel.objects.filter(pattern__startswith=str(request.user.id))
    journal_recieved_number = request.POST.get('journal_no')
    print(f'================== journal_recieved_number = {journal_recieved_number}==================')
    if journals.exists():
        last = journals.last()
        last_id = last.jn_rec_number
        print(f'================== last_id = {last_id}==================')
        if journal_recieved_number == last_id:
            return True
        else:
            return False
    else:
        # if payments_recieved_number != 'PRN-01':
        #     return HttpResponse("<span class='text-danger'>Payment Recieved Number is not Continues</span>")
        # else:
        #     return HttpResponse("")
        return True

def journal(request):
    return render(request,'zohomodules/manual_journal/add_journal.html')



def add_journal(request):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    
        log_details = LoginDetails.objects.get(id=log_id)
        print("hloo")
        print(log_details)

        if log_details.user_type == "Company":
            dash_details = CompanyDetails.objects.get(login_details=log_details)
            print("a")
            print(dash_details)
            allmodules = ZohoModules.objects.get(company=dash_details, status='New')
            print("b")
            print(allmodules)
            accounts = Chart_of_Accounts.objects.filter(company=dash_details)
            print("c")
            print(accounts)
            journal_no = request.POST.get('journal_no')
            reference_no = ''
            jon = JournalRecievedIdModel.objects.filter(company=dash_details)
            last = ''
            if jon.exists():
                last = jon.last()


            if request.method == 'POST':
                
                user = request.user
                date = request.POST.get('date')
                journal_no = request.POST.get('journal_no')
                reference_no=request.POST.get('reference_no')

                notes = request.POST.get('notes')
                currency = request.POST.get('currency')
                cash_journal = request.POST.get('cash_journal') == 'True'

                attachment = request.FILES.get('attachment')
                print("o")
                print(attachment)
                status = ""  # Default value for status
                if 'Draft' in request.POST:
                    status="draft"
                if "Save" in request.POST:
                    status="save"  

                journal = Journal(
                    date=date,
                    journal_no=journal_no,
                    reference_no=reference_no,
                    notes=notes,
                    currency=currency,
                    journal_type=cash_journal,
                    attachment=attachment,
                    status=status,
                    company=dash_details if log_details.user_type == "Company" else None   
                )
                if attachment:
                    # File was uploaded, proceed with saving it
                    journal.attachment = attachment
                journal.save()
                
                JournalTransactionHistory.objects.create(
                    company=dash_details  if log_details.user_type == "Company" else None,
                    login_details=log_details,
                    journal=journal,
                    action='Created',
                )

                account_list = request.POST.getlist('account')
                description_list = request.POST.getlist('description')
                contact_list = request.POST.getlist('contact')
                debits_list = request.POST.getlist('debits')
                credits_list = request.POST.getlist('credits')

                total_debit = 0
                total_credit = 0

                for i in range(len(account_list)):
                    account = account_list[i]
                    description = description_list[i]
                    contact = contact_list[i]
                    debits = debits_list[i]
                    credits = credits_list[i]

                    journal_entry = JournalEntry(
                        journal=journal,
                        account=account,
                        description=description,
                        contact=contact,
                        debits=debits,
                        credits=credits
                    )
                    journal_entry.save()

                    total_debit += float(debits) if debits else 0
                    total_credit += float(credits) if credits else 0

                debit_difference = total_debit - total_credit
                credit_difference = total_credit - total_debit

                journal.total_debit = total_debit
                journal.total_credit = total_credit
                journal.debit_difference = debit_difference
                journal.credit_difference=credit_difference
                journal_no = request.POST.get('journal_no')    
                reference_no = request.POST.get('reference_no')
                journal.reference_no=reference_no
                print(reference_no)
                journal.save()
        
                is_valid = check_journal_num_valid(request)
                print("good")
                print(is_valid)
                if not is_valid:
                    messages.error(request, 'Invalid journal number. Please enter a valid and continuous numeric sequence.')

                if JournalRecievedIdModel.objects.filter(company=dash_details).exists():
                    jn = JournalRecievedIdModel.objects.filter(company=dash_details)
                    jn_id = jn.last()
                    jn_id1 = jn.last()

                    # Check if there is a second last journal record
                    if jn.exclude(id=jn_id.id).last():
                        jn_id_second_last = jn.exclude(id=jn_id.id).last()
                        pattern = jn_id_second_last.pattern
                    else:
                        jn_id_second_last = jn.first()
                        pattern = jn_id_second_last.pattern

                    if journal_no != jn_id.jn_rec_number and journal_no != '':
                        # Creating a new JournalRecievedIdModel instance
                        jn_id = JournalRecievedIdModel(company=dash_details)
                        count_for_ref_no = JournalRecievedIdModel.objects.filter(company=dash_details.id).count()
                        jn_id.pattern = pattern
                        jn_id.save()

                        # Using count_for_ref_no + 1 as the reference number
                        ref_num = int(count_for_ref_no) + 2
                        jn_id.ref_number = f'{ref_num:02}'

                        jn_id.jn_rec_number = jn_id1.jn_rec_number
                        jn_id.save()
                    else:
                        # Creating a new JournalRecievedIdModel instance
                        jn_id = JournalRecievedIdModel(company=dash_details)
                        count_for_ref_no = JournalRecievedIdModel.objects.filter(company=dash_details.id).count()
                        jn_id.pattern = pattern
                        jn_id.save()

                        # Using count_for_ref_no + 1 as the reference number
                        ref_num = int(count_for_ref_no) + 2
                        jn_id.ref_number = f'{ref_num:02}'

                        # Incrementing the jn_rec_number
                        jn_rec_num = ''.join(i for i in jn_id1.jn_rec_number if i.isdigit())
                        jn_rec_num = int(jn_rec_num)+1
                        print("#################################")
                        print(f"-----------------{jn_id1}-----------------")
                        jn_id.jn_rec_number = f'{pattern}{jn_rec_num:02}'
                        print(jn_id.jn_rec_number)
                        jn_id.save()
                        
                else:
                    # Creating a new JournalRecievedIdModel instance
                    jn_id = JournalRecievedIdModel(company=dash_details)
                    jn_id.save()

                    # Setting initial values for ref_number, pattern, and jn_rec_number
                    jn_id.ref_number = f'{2:02}'

                    pattern = ''.join(i for i in journal_no if not i.isdigit())
                    jn_id.pattern = pattern
                    jn_id.jn_rec_number = f'{pattern}{2:02}'
                    jn_id.save()
            
                if not is_valid:
                    return redirect('add_journal')
                else:
                    return redirect('manual_journal')
            context = {
                 'log_id':log_id,
                 'log_details':log_details,
                'details': dash_details,
                'allmodules': allmodules,
                'reference_no': reference_no,
                 'last':last,
                 'accounts':accounts,
            }
            return render(request, 'zohomodules/manual_journal/add_journal.html',context)

        elif log_details.user_type == 'Staff':
            company_details = StaffDetails.objects.get(login_details=log_details)
            print("c")
            print(company_details)
            comp=CompanyDetails.objects.get(id=company_details.company.id)
            print("d")
            print(comp)
            allmodules = ZohoModules.objects.get(company=company_details.company, status='New')
            print("e")
            print(allmodules)
            jour = JournalRecievedIdModel.objects.filter(staff=company_details)
            accounts = Chart_of_Accounts.objects.filter(company=company_details.company)
        
            journal_no = request.POST.get('journal_no')
            reference_no = ''
            jon = JournalRecievedIdModel.objects.filter(staff=company_details)
            last = ''
            if jon.exists():
                last = jon.last()


            if request.method == 'POST':
                
                user = request.user
                date = request.POST.get('date')
                journal_no = request.POST.get('journal_no')
                reference_no=request.POST.get('reference_no')
                notes = request.POST.get('notes')
                currency = request.POST.get('currency')
                cash_journal = request.POST.get('cash_journal') == 'True'

                attachment = request.FILES.get('attachment')
                print("o")
                print(attachment)
                
                status = ""  # Default value for status
             
                if 'Draft' in request.POST:
                    status="draft"
                if "Save" in request.POST:
                    status="save"
                journal = Journal(
                    date=date,
                    journal_no=journal_no,
                    reference_no=reference_no,
                    notes=notes,
                    currency=currency,
                    journal_type=cash_journal,
                    attachment=attachment, 
                    status=status,
                    staff=company_details if log_details.user_type == 'Staff' else None 
                )
                if attachment:
                    # File was uploaded, proceed with saving it
                    journal.attachment = attachment
                journal.save()
                
                JournalTransactionHistory.objects.create(
                    staff=company_details if log_details.user_type == 'Staff' else None,
                    login_details=log_details,
                    journal=journal,
                    action='Created',
                )
            
                
                account_list = request.POST.getlist('account')
                description_list = request.POST.getlist('description')
                contact_list = request.POST.getlist('contact')
                debits_list = request.POST.getlist('debits')
                credits_list = request.POST.getlist('credits')

                total_debit = 0
                total_credit = 0

                for i in range(len(account_list)):
                    account = account_list[i]
                    description = description_list[i]
                    contact = contact_list[i]
                    debits = debits_list[i]
                    credits = credits_list[i]

                    journal_entry = JournalEntry(
                        journal=journal,
                        account=account,
                        description=description,
                        contact=contact,
                        debits=debits,
                        credits=credits
                    )
                    journal_entry.save()

                    total_debit += float(debits) if debits else 0
                    total_credit += float(credits) if credits else 0

                debit_difference = total_debit - total_credit
                credit_difference = total_credit - total_debit

                journal.total_debit = total_debit
                journal.total_credit = total_credit
                journal.debit_difference = debit_difference
                journal.credit_difference=credit_difference
                journal_no = request.POST.get('journal_no')    
                reference_no = request.POST.get('reference_no')
                journal.reference_no=reference_no
                print(reference_no)
                journal.save()
        
                is_valid = check_journal_num_valid(request)
                print(is_valid)
                if not is_valid:
                    messages.error(request, 'Invalid journal number. Please enter a valid and continuous numeric sequence.')

                if JournalRecievedIdModel.objects.filter(staff=company_details).exists():
                    jn = JournalRecievedIdModel.objects.filter(staff=company_details)
                    jn_id = jn.last()
                    jn_id1 = jn.last()

                    # Check if there is a second last journal record
                    if jn.exclude(id=jn_id.id).last():
                        jn_id_second_last = jn.exclude(id=jn_id.id).last()
                        pattern = jn_id_second_last.pattern
                    else:
                        jn_id_second_last = jn.first()
                        pattern = jn_id_second_last.pattern

                    if journal_no != jn_id.jn_rec_number and journal_no != '':
                        # Creating a new JournalRecievedIdModel instance
                        jn_id = JournalRecievedIdModel(staff=company_details)
                        count_for_ref_no = JournalRecievedIdModel.objects.filter(staff=company_details.id).count()
                        jn_id.pattern = pattern
                        jn_id.save()

                        # Using count_for_ref_no + 1 as the reference number
                        ref_num = int(count_for_ref_no) + 2
                        jn_id.ref_number = f'{ref_num:02}'

                        jn_id.jn_rec_number = jn_id1.jn_rec_number
                        jn_id.save()
                    else:
                        # Creating a new JournalRecievedIdModel instance
                        jn_id = JournalRecievedIdModel(staff=company_details)
                        count_for_ref_no = JournalRecievedIdModel.objects.filter(staff=company_details.id).count()
                        jn_id.pattern = pattern
                        jn_id.save()

                        # Using count_for_ref_no + 1 as the reference number
                        ref_num = int(count_for_ref_no) + 2
                        jn_id.ref_number = f'{ref_num:02}'

                        # Incrementing the jn_rec_number
                        jn_rec_num = ''.join(i for i in jn_id1.jn_rec_number if i.isdigit())
                        jn_rec_num = int(jn_rec_num)+1
                        print("#################################")
                        print(f"-----------------{jn_id1}-----------------")
                        jn_id.jn_rec_number = f'{pattern}{jn_rec_num:02}'
                        print(jn_id.jn_rec_number)
                        jn_id.save()
                        
                else:
                    # Creating a new JournalRecievedIdModel instance
                    jn_id = JournalRecievedIdModel(staff=company_details)
                    jn_id.save()

                    # Setting initial values for ref_number, pattern, and jn_rec_number
                    jn_id.ref_number = f'{2:02}'

                    pattern = ''.join(i for i in journal_no if not i.isdigit())
                    jn_id.pattern = pattern
                    jn_id.jn_rec_number = f'{pattern}{2:02}'
                    jn_id.save()
            
                if not is_valid:
                    return redirect('add_journal')
                else:
                    return redirect('manual_journal')
                
            context = {
                 'log_id':log_id,
                 'log_details':log_details,
                'details':company_details,
                'allmodules': allmodules,
                'reference_no': reference_no,
                 'last':last,
                 'jour':jour,
                 'accounts':accounts,
                 'company':comp,
            }

    return render(request, 'zohomodules/manual_journal/add_journal.html',context)



def journal_overview(request, journal_id):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    
    if log_details.user_type=="Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        journal = Journal.objects.filter(company=dash_details)
        jour = get_object_or_404(Journal, id=journal_id)
        #journal_entries = JournalEntry.objects.filter(journal__in=journal,company=dash_details)
        journal_entries = JournalEntry.objects.filter(journal=jour)
        
        comments = JournalComment.objects.filter(journal=jour)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        sort_option = request.GET.get('sort', 'all')  
        filter_option = request.GET.get('filter', 'all')
        if sort_option == 'journal_no':
            journal = journal.order_by('journal_no')
        elif sort_option == 'total_debit':
            journal = journal.order_by('total_debit')

        if filter_option == 'save':
            journal = journal.filter(status='save')
        elif filter_option == 'draft':
            journal = journal.filter(status='draft')
        transaction_history = JournalTransactionHistory.objects.filter(journal=jour)
        #items = PriceListItem.objects.filter(company=dash_details, price_list=price_list)
        latest_transaction =JournalTransactionHistory.objects.filter(journal=jour)

        context={
            'log_id':log_id,
            'log_details':log_details,
            'details':dash_details,
            'allmodules': allmodules,
            'journal': journal,
            'jour': jour,
            'journal_entries':journal_entries,
            'comments': comments,
            'sort_option': sort_option,
            'filter_option': filter_option,
            'latest_transaction': latest_transaction,
            'transaction_history': transaction_history,
            #'items':items,
        }
        return render(request,'zohomodules/manual_journal/journal_list.html',context)
    
    if log_details.user_type=="Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        journal = Journal.objects.filter(staff=dash_details)
        jour = get_object_or_404(Journal, id=journal_id)
        journal_entries = JournalEntry.objects.filter(journal=jour)
        comments = JournalComment.objects.filter(journal=jour)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        sort_option = request.GET.get('sort', 'all')  
        filter_option = request.GET.get('filter', 'all')
        if sort_option == 'journal_no':
            journal = journal.order_by('journal_no')
        elif sort_option == 'total_debit':
            journal = journal.order_by('total_debit')

        if filter_option == 'save':
            journal = journal.filter(status='save')
        elif filter_option == 'draft':
            journal = journal.filter(status='draft')
        transaction_history = JournalTransactionHistory.objects.filter(journal=jour)
        #items = PriceListItem.objects.filter(company=dash_details.company, price_list=price_list)
        context={
            'log_id':log_id,
            'log_details':log_details,
            'details':dash_details,
            'allmodules': allmodules,
            'journal': journal,
            'comments': comments,
            'jour': jour,
            'journal_entries':journal_entries,
            'sort_option': sort_option,
            'filter_option': filter_option,
            'transaction_history': transaction_history,
            #'items':items,
        }
        return render(request,'zohomodules/manual_journal/journal_list.html',context)
    
    
    

def update_journal_status(request,id):
    jo=Journal.objects.get(id=id)
    jo.status = "save"
    jo.save()
    return redirect('journal_overview', id)


def delete_journal(request, journal_id):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details= LoginDetails.objects.get(id=log_id)
    if log_details.user_type=="Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        
        journal = Journal.objects.filter(company=dash_details)
        jour = get_object_or_404(Journal, id=journal_id)
        journal_entries = JournalEntry.objects.filter(journal=jour)
        allmodules= ZohoModules.objects.get(company=dash_details,status='New')
        jour.delete()
        context={
            'details':dash_details,
            'allmodules': allmodules,
            'journal': journal,
            'jour': jour,
            'journal_entries':journal_entries,
        }
        return render(request,'zohomodules/manual_journal/manual_journal.html',context)
    if log_details.user_type=="Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        journal = Journal.objects.filter(staff=dash_details)
        jour = get_object_or_404(Journal, id=journal_id)
        journal_entries = JournalEntry.objects.filter(journal=jour)
        allmodules= ZohoModules.objects.get(company=dash_details.company,status='New')
        jour.delete()
        context={
            'details':dash_details,
            'allmodules': allmodules,
            'journal': journal,
            'jour': jour,
            'journal_entries':journal_entries,
        }
        return render(request,'zohomodules/manual_journal/manual_journal.html',context)


def add_journal_comment(request, journal_id):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details = LoginDetails.objects.get(id=log_id)
    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        journal = get_object_or_404(Journal, id=journal_id, company=dash_details)
        if request.method == 'POST':
            comment = request.POST.get('comment_text')
            JournalComment.objects.create(
                company=dash_details,
                login_details=log_details,
                journal=journal,
                comment=comment
            )
            
        return redirect('journal_overview', journal_id=journal_id)
    if log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        journal = get_object_or_404(Journal, id=journal_id, staff=dash_details)
        if request.method == 'POST':
            comment = request.POST.get('comment_text')
            JournalComment.objects.create(
                staff=dash_details,
                login_details=log_details,
                journal=journal,
                comment=comment
            )
        return redirect('journal_overview', journal_id=journal_id)

def delete_journal_comment(request, comment_id, journal_id):
    comment = get_object_or_404(JournalComment, id=comment_id, journal_id=journal_id)
    comment.delete()
    return redirect('journal_overview', journal_id=journal_id)


def create_account_jour(request):                                                                #new by tinto mt
    login_id = request.session['login_id']
    log_user = LoginDetails.objects.get(id=login_id)
    if log_user.user_type == 'Company':
        company_id = request.session['login_id']
        if request.method=='POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            account=Chart_of_Accounts.objects.all()
            c = CompanyDetails.objects.get(login_details=company_id)
            b.company=c
            b.logindetails=log_user
            b.action="Created"
            b.Date=date.today()
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.account_number = request.POST.get("account_number2",None)
            a.account_description = request.POST['description']
            if a.account_type=="Other Current Assets":

                a.sub_account = request.POST.get("sub_account",None)
                a.parent_account = request.POST.get("parent_account",None)
               

            if a.account_type=="Cash":
             
                a.sub_account = request.POST.get("sub_account22",None)
                a.parent_account = request.POST.get("parent_account22",None)
               

            if a.account_type=="Fixed Assets":
            
                a.sub_account = request.POST.get("sub_account33",None)
                a.parent_account = request.POST.get("parent_account33",None)
               
            
            if a.account_type=="Stock":
               
                a.sub_account = request.POST.get("sub_account44",None)
                a.parent_account = request.POST.get("parent_account44",None)
             
            
            if a.account_type=="Other Current Liability":
             
                a.sub_account = request.POST.get("sub_account55",None)
                a.parent_account = request.POST.get("parent_account55",None)
               
            if a.account_type=="Long Term Liability":
            
                a.sub_account = request.POST.get("sub_account66",None)
                a.parent_account = request.POST.get("parent_account66",None)
              
            
            if a.account_type=="Other Liability":
              
                a.sub_account = request.POST.get("sub_account77",None)
                a.parent_account = request.POST.get("parent_account77",None)
              
            if a.account_type=="Equity":
            
                a.sub_account = request.POST.get("sub_account88",None)
                a.parent_account = request.POST.get("parent_account88",None)
            
            
            if a.account_type=="Income":
             
                a.sub_account = request.POST.get("sub_account99",None)
                a.parent_account = request.POST.get("parent_account99",None)
              
            
            if a.account_type=="Expense":
             
                a.sub_account = request.POST.get("sub_account100",None)
                a.parent_account = request.POST.get("parent_account100",None)
              
            if a.account_type=="Cost Of Goods Sold":
              
                a.sub_account = request.POST.get("sub_account111",None)
                a.parent_account = request.POST.get("parent_account111",None)
             
            if a.account_type=="Other Expense":
             
                a.sub_account = request.POST.get("sub_account222",None)
                a.parent_account = request.POST.get("parent_account222",None)
               
            account_type=request.POST.get("account_type",None)
            if account_type == 'Other Assets':
                a.description = 'Track special assets like goodwill and other intangible assets'
            if account_type == 'Other Current Assets':
                a.description = 'Any short term asset that can be converted into cash or cash equivalents easily Prepaid expenses Stocks and Mutual Funds'
            if account_type == 'Cash':
                a.description = 'To keep track of cash and other cash equivalents like petty cash, undeposited funds, etc., use an organized accounting system  financial software'
            if account_type == 'Bank':
                a.description = 'To keep track of bank accounts like Savings, Checking, and Money Market accounts.'
            if account_type == 'Fixed Asset':
                a.description = 'Any long-term investment or asset that cannot be easily converted into cash includes: Land and Buildings, Plant, Machinery, and Equipment, Computers, Furniture.'
            if account_type == 'Stock':
                a.description = 'To keep track of your inventory assets.'
            if account_type == 'Payment Clearing':
                a.description = 'To keep track of funds moving in and out via payment processors like Stripe, PayPal, etc.'
            if account_type == 'Other Liability':
                a.description = 'Obligation of an entity arising from past transactions or events which would require repayment.Tax to be paid Loan to be Repaid Accounts Payableetc.'
            if account_type == 'Other Current Liability':
                a.description = 'Any short term liability like: Customer Deposits Tax Payable'
            if account_type == 'Credit Card':
                a.description = 'Create a trail of all your credit card transactions by creating a credit card account.'
            if account_type == 'Long Term Liability':
                a.description = 'Liabilities that mature after a minimum period of one year like: Notes Payable Debentures Long Term Loans '
            if account_type == 'Overseas Tax Payable':
                a.description = 'Track your taxes in this account if your business sells digital services to foreign customers.'
            if account_type == 'Equity':
                a.description = 'Owners or stakeholders interest on the assets of the business after deducting all the liabilities.'
            if account_type == 'Income':
                a.description = 'Income or Revenue earned from normal business activities like sale of goods and services to customers.'
            if account_type == 'Other Income':
                a.description = 'Income or revenue earned from activities not directly related to your business like : Interest Earned Dividend Earned'
            if account_type == 'Expense':
                a.description = 'Reflects expenses incurred for running normal business operations, such as : Advertisements and Marketing Business Travel Expenses License Fees Utility Expenses'
            if account_type == 'Cost Of Goods Sold':
                a.description = 'This indicates the direct costs attributable to the production of the goods sold by a company such as: Material and Labor costs Cost of obtaining raw materials'
            if account_type == 'Other Expense':
                a.description = 'Track miscellaneous expenses incurred for activities other than primary business operations or create additional accounts to track default expenses like insurance or contribution towards charity.'
       

            
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name,company=c).exists():
                error='yes'
                messages.error(request,'Account with same name exsits !!!')
                return redirect('add_journal')
            else:
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                return redirect('add_journal')
    elif log_user.user_type == 'Staff':
        staff_id = request.session['login_id']
        if request.method=='POST':
            a=Chart_of_Accounts()
            b=Chart_of_Accounts_History()
            staff = LoginDetails.objects.get(id=staff_id)
            sf = StaffDetails.objects.get(login_details=staff)
            c=sf.company
            b.Date=date.today()
            b.company=c
            b.logindetails=log_user
            a.login_details=log_user
            a.company=c
          
        
            a.account_type = request.POST.get("account_type",None)
            a.account_name = request.POST.get("account_name",None)
            a.account_code = request.POST.get("account_code",None)
            a.account_number = request.POST.get("account_number2",None)
            a.account_description = request.POST['description']
            account_type=request.POST.get("account_type",None)
            if a.account_type=="Other Current Assets":

                a.sub_account = request.POST.get("sub_account",None)
                a.parent_account = request.POST.get("parent_account",None)
               

            if a.account_type=="Cash":
             
                a.sub_account = request.POST.get("sub_account22",None)
                a.parent_account = request.POST.get("parent_account22",None)
               

            if a.account_type=="Fixed Assets":
            
                a.sub_account = request.POST.get("sub_account33",None)
                a.parent_account = request.POST.get("parent_account33",None)
               
            
            if a.account_type=="Stock":
               
                a.sub_account = request.POST.get("sub_account44",None)
                a.parent_account = request.POST.get("parent_account44",None)
             
            
            if a.account_type=="Other Current Liability":
             
                a.sub_account = request.POST.get("sub_account55",None)
                a.parent_account = request.POST.get("parent_account55",None)
               
            if a.account_type=="Long Term Liability":
            
                a.sub_account = request.POST.get("sub_account66",None)
                a.parent_account = request.POST.get("parent_account66",None)
              
            
            if a.account_type=="Other Liability":
              
                a.sub_account = request.POST.get("sub_account77",None)
                a.parent_account = request.POST.get("parent_account77",None)
              
            if a.account_type=="Equity":
            
                a.sub_account = request.POST.get("sub_account88",None)
                a.parent_account = request.POST.get("parent_account88",None)
            
            
            if a.account_type=="Income":
             
                a.sub_account = request.POST.get("sub_account99",None)
                a.parent_account = request.POST.get("parent_account99",None)
              
            
            if a.account_type=="Expense":
             
                a.sub_account = request.POST.get("sub_account100",None)
                a.parent_account = request.POST.get("parent_account100",None)
              
            if a.account_type=="Cost Of Goods Sold":
              
                a.sub_account = request.POST.get("sub_account111",None)
                a.parent_account = request.POST.get("parent_account111",None)
             
            if a.account_type=="Other Expense":
             
                a.sub_account = request.POST.get("sub_account222",None)
                a.parent_account = request.POST.get("parent_account222",None)
               
            account_type=request.POST.get("account_type",None)
            if account_type == 'Other Assets':
                a.description = 'Track special assets like goodwill and other intangible assets'
            if account_type == 'Other Current Assets':
                a.description = 'Any short term asset that can be converted into cash or cash equivalents easily Prepaid expenses Stocks and Mutual Funds'
            if account_type == 'Cash':
                a.description = 'To keep track of cash and other cash equivalents like petty cash, undeposited funds, etc., use an organized accounting system  financial software'
            if account_type == 'Bank':
                a.description = 'To keep track of bank accounts like Savings, Checking, and Money Market accounts.'
            if account_type == 'Fixed Asset':
                a.description = 'Any long-term investment or asset that cannot be easily converted into cash includes: Land and Buildings, Plant, Machinery, and Equipment, Computers, Furniture.'
            if account_type == 'Stock':
                a.description = 'To keep track of your inventory assets.'
            if account_type == 'Payment Clearing':
                a.description = 'To keep track of funds moving in and out via payment processors like Stripe, PayPal, etc.'
            if account_type == 'Other Liability':
                a.description = 'Obligation of an entity arising from past transactions or events which would require repayment.Tax to be paid Loan to be Repaid Accounts Payableetc.'
            if account_type == 'Other Current Liability':
                a.description = 'Any short term liability like: Customer Deposits Tax Payable'
            if account_type == 'Credit Card':
                a.description = 'Create a trail of all your credit card transactions by creating a credit card account.'
            if account_type == 'Long Term Liability':
                a.description = 'Liabilities that mature after a minimum period of one year like: Notes Payable Debentures Long Term Loans '
            if account_type == 'Overseas Tax Payable':
                a.description = 'Track your taxes in this account if your business sells digital services to foreign customers.'
            if account_type == 'Equity':
                a.description = 'Owners or stakeholders interest on the assets of the business after deducting all the liabilities.'
            if account_type == 'Income':
                a.description = 'Income or Revenue earned from normal business activities like sale of goods and services to customers.'
            if account_type == 'Other Income':
                a.description = 'Income or revenue earned from activities not directly related to your business like : Interest Earned Dividend Earned'
            if account_type == 'Expense':
                a.description = 'Reflects expenses incurred for running normal business operations, such as : Advertisements and Marketing Business Travel Expenses License Fees Utility Expenses'
            if account_type == 'Cost Of Goods Sold':
                a.description = 'This indicates the direct costs attributable to the production of the goods sold by a company such as: Material and Labor costs Cost of obtaining raw materials'
            if account_type == 'Other Expense':
                a.description = 'Track miscellaneous expenses incurred for activities other than primary business operations or create additional accounts to track default expenses like insurance or contribution towards charity.'
       
    
            a.Create_status="active"
            ac_name=request.POST.get("account_name",None)
            if Chart_of_Accounts.objects.filter(account_name=ac_name,company=c).exists():
                error='yes'
                messages.error(request,'Account with same name exsits')
                return redirect('add_journal')
            else:
                a.save()
                t=Chart_of_Accounts.objects.get(id=a.id)
                b.chart_of_accounts=t
                b.save()
                return redirect('add_journal')

    return redirect('add_journal')

def edit_journal(request, journal_id):
    if 'login_id' in request.session:
        if request.session.has_key('login_id'):
            log_id = request.session['login_id']
        else:
            return redirect('/')
    log_details = LoginDetails.objects.get(id=log_id)
    if log_details.user_type == "Company":
        dash_details = CompanyDetails.objects.get(login_details=log_details)
        allmodules = ZohoModules.objects.get(company=dash_details, status='New')
        journal = get_object_or_404(Journal, id=journal_id, company=dash_details)
        journal_entries = JournalEntry.objects.filter(journal=journal)


        
        if request.method == 'POST':
            date = request.POST.get('date')
            journal_no = request.POST.get('journal_no')
            #reference_no = request.POST.get('reference_no')
            notes = request.POST.get('notes')
            currency = request.POST.get('currency')
            cash_journal = request.POST.get('cash_journal') == 'True'

            journal.date = date
            journal.journal_no = journal_no
            #journal.reference_no = reference_no
            journal.notes = notes
            journal.currency = currency
            journal.cash_journal = cash_journal       
            journal.user = request.user
            old=journal.attachment
            new = request.FILES.get('attachment')
            if old !=None and new==None:
                journal.attachment=old
            else:
                journal.attachment=new            
            journal.save()

            account_list = request.POST.getlist('account')
            description_list = request.POST.getlist('description')
            contact_list = request.POST.getlist('contact')
            debits_list = request.POST.getlist('debits')
            credits_list = request.POST.getlist('credits')

            total_debit = 0
            total_credit = 0

            JournalEntry.objects.filter(journal=journal).delete()

            for i in range(len(account_list)):
                account = account_list[i]
                description = description_list[i]
                contact = contact_list[i]
                debits = debits_list[i]
                credits = credits_list[i]

                journal_entry = JournalEntry(
                    journal=journal,
                    account=account,
                    description=description,
                    contact=contact,
                    debits=debits,
                    credits=credits
                )
                journal_entry.save()

                total_debit += float(debits) if debits else 0
                total_credit += float(credits) if credits else 0

            difference = total_debit - total_credit

            journal.total_debit = total_debit
            journal.total_credit = total_credit
            journal.difference = difference
            journal.save()
            JournalTransactionHistory.objects.create(
                        company=dash_details,
                        login_details=log_details,
                        journal=journal,
                        action='Edited',
                    )

            return redirect('journal_overview', journal_id=journal_id)
        context = {
            'log_details': log_details,
            'details': dash_details,
            'allmodules': allmodules,
            'journal': journal,
            'journal_entries': journal_entries,
            
        }
        return render(request, 'zohomodules/manual_journal/edit_journal.html', context)
    elif log_details.user_type == "Staff":
        dash_details = StaffDetails.objects.get(login_details=log_details)
        allmodules = ZohoModules.objects.get(company=dash_details.company, status='New')
        journal = get_object_or_404(Journal, id=journal_id, company=dash_details.company)
        journal_entries = JournalEntry.objects.filter(journal=journal)
        accounts = Chart_of_Accounts.objects.filter(company=dash_details.company)
        
        if request.method == 'POST':
            date = request.POST.get('date')
            journal_no = request.POST.get('journal_no')
            #reference_no = request.POST.get('reference_no')
            notes = request.POST.get('notes')
            currency = request.POST.get('currency')
            cash_journal = request.POST.get('cash_journal') == 'True'

            journal.date = date
            journal.journal_no = journal_no
            #journal.reference_no = reference_no
            journal.notes = notes
            journal.currency = currency
            journal.journal_type = cash_journal       
            journal.user = request.user
            # Handle attachment
            old_attachment = journal.attachment
            new_attachment = request.FILES.get('attachment')
            if old_attachment and not new_attachment:
                journal.attachment = old_attachment
            elif new_attachment:
                journal.attachment = new_attachment
            # Save the changes to the Journal object
            journal.save()

            account_list = request.POST.getlist('account')
            description_list = request.POST.getlist('description')
            contact_list = request.POST.getlist('contact')
            debits_list = request.POST.getlist('debits')
            credits_list = request.POST.getlist('credits')

            total_debit = 0
            total_credit = 0

            JournalEntry.objects.filter(journal=journal).delete()

            for i in range(len(account_list)):
                account = account_list[i]
                description = description_list[i]
                contact = contact_list[i]
                debits = debits_list[i]
                credits = credits_list[i]

                journal_entry = JournalEntry(
                    journal=journal,
                    account=account,
                    description=description,
                    contact=contact,
                    debits=debits,
                    credits=credits
                )
                journal_entry.save()

                total_debit += float(debits) if debits else 0
                total_credit += float(credits) if credits else 0

            difference = total_debit - total_credit

            journal.total_debit = total_debit
            journal.total_credit = total_credit
            journal.difference = difference
            journal.save()
            JournalTransactionHistory.objects.create(
                        company=dash_details,
                        login_details=log_details,
                        journal=journal,
                        action='Edited',
                    )
            
            return redirect('journal_overview', journal_id=journal_id)

        context = {
            'details': dash_details,
            'allmodules': allmodules,
            'journal': journal,
            'journal_entries': journal_entries,
            'accounts':accounts,
        }
        return render(request, 'zohomodules/manual_journal/edit_journal.html', context)








